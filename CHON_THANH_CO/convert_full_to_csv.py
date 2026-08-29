#!/usr/bin/env python3
import openpyxl
import csv
import os

wb = openpyxl.load_workbook('san_pham_dia_ky_thuat_chi_tiet.xlsx')
output_dir = 'csv_export_full'
os.makedirs(output_dir, exist_ok=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Create filename from sheet name
    safe_name = sheet_name.lower().replace(' ', '_').replace('(', '').replace(')', '')
    filename = f"{output_dir}/{safe_name}.csv"
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(cell) if cell is not None else '' for cell in row])
    print(f"Exported: {filename} ({ws.max_row} rows)")
