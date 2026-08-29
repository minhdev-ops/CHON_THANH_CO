#!/usr/bin/env python3
"""
Tạo file Excel tổng hợp chi tiết sản phẩm từ 3 website địa kỹ thuật Việt Nam.
thaichau.vn | bactham.vn | aritex.com.vn
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
cat_font = Font(name='Arial', bold=True, size=11, color='1F4E79')
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='top', wrap_text=True)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def setup_sheet(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def write_data(ws, data, start_row=2):
    for r, row in enumerate(data, start_row):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap
            cell.border = border
            cell.font = Font(name='Arial', size=10)

# ═══════════════════════════════════════════════════════
# SHEET 1: DANH MỤC SẢN PHẨM
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Danh muc san pham"

h1 = ["STT", "Nha cung cap", "Website", "Nhom san pham", "Ten san pham", "Ma san pham",
      "Mo ta chi tiet", "Dac diem noi bat", "Chuc nang", "Ung dung", "Tieu chuan"]
w1 = [5, 25, 20, 25, 35, 20, 65, 50, 45, 50, 25]
setup_sheet(ws1, h1, w1)

products = [
    # ═══════════════════════════════════════════════════════
    # THAI CHAU - thaichau.vn
    # ═══════════════════════════════════════════════════════
    [1, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "Vai dia ky thuat khong det",
     "Vai dia ky thuat khong det APT",
     "APT 7 - APT 80",
     "Vai dia ky thuat khong det APT duoc san xuat hoan toan tren day chuyen tu dong voi cong nghe Han Quoc, su dung 100% nguyen lieu polypropylene, xuyen kim co phu gia khang tia cuc tlim. Nha may APT co the san xuat tu cuong luc 7kN den 80kN (100g/m2 den 1200g/m2). Vai dia ky thuat khong det APT duoc dung rong rai trong cac cong trinh xay dung cau duong nho chat luong cao, tinh nang da dang phu hop voi moi yeu cau ky thuat.",
     "1. San pham da dang, dap ung duoc moi yeu cau ky thuat\n2. San pham co tinh on dinh cao\n3. Gia thanh hop ly voi chat luong quoc te\n4. Nguon nguyen lieu chat luong cao, nhap khau tu nuoc ngoai\n5. Ho tro ky thuat tu doi ngu ky su chuyen mon cao",
     "1. Phan cach: Tiet kiem chi phi xay dung, ngan chan mat vat lieu, giam khoi luong dat dap\n2. Loc nguoc: Tranh xi mon tu vat lieu hat min vao vat lieu tho\n3. Tieu thoat: Thu tieu tan nuoc trong lo rom nhanh, giam ap luc nuoc lo rom\n4. Bao ve: Ket hop voi ro da, da hoc, be tong tao lop dem chong xi\n5. Gia cuong: Truyen cho dat cuong do chiu keo, chong pha hoai cuc bo",
     "1. On dinh nen duong bo, duong sat, san kho, bai Container\n2. Khoi phuc va gia cuong nen dat yeu (dam, ao bun)\n3. Chong xi mon de, dap, kenh muong thuy loi, ke song, bien\n4. Loc va thoat nuoc nen duong, san van dong, san golf, cong vien\n5. Bao ve mang chong tham trong he thong xu ly chat thai, ho chua nuoc thai",
     "ASTM D4595 / TCVN 8485"],

    [2, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "Vai dia ky thuat det",
     "Vai dia ky thuat det DML",
     "DML",
     "Vai dia ky thuat det DML do nha may APT san xuat tren day chuyen tu dong theo tieu chuan cong nghe Han Quoc, su dung 100% nguyen lieu tu Polyester. Vai dia ky thuat det DML duoc su dung o cac cong trinh trong diem nhu cau can, duong lon, be ke quan trong nho co cuong do chiu keo cao va do dan thap. Vai det DML dam bao yeu cau tuoi tho dai han va do ben cho cac cong trinh.",
     "1. Cuong do chiu keo cao\n2. Do dan thap, on dinh hinh dang\n3. Kha nang loc va thoat nuoc tot\n4. Do ben moi truong cao\n5. Phu hop cong trinh trong diem",
     "1. On dinh va gia cuong nen dat yeu: Chia nho suc ep cua lop dat tho, ngan khong cho dat bi dao thanh hoc nho\n2. Loc va thoat nuoc: He so thấm, toc do dong chay cao, giu lai hat dat khong bi lap tac\n3. Phan cach va on dinh muong ranh: Do ben huong len giu chat duong on, la lop phan cach\n4. Chong xi mon: Giu cho dat khong bi xi mon truoc su tan cong cua song bien",
     "1. Khoi phuc nen dat yeu (dam, ao bun)\n2. Lien ket cac coc gia co nen (nha xuong, bon, be)\n3. Dem nen co nhieu lo trong (nen da vloi)\n4. Chong xi mon - loc va tieu thoat (de, dap, kenh muong, ke song, bien)\n5. Gia co nen duong dap tren nen dat yeu",
     "ASTM D4595 / TCVN 8485"],

    [3, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "Bac tham",
     "Bac tham PVD",
     "VID / RID",
     "Bac tham la vat lieu dia ky thuat dung de thoat nuoc muc dich gia tang qua trinh co ket cua nen mong. Cau tao gom 2 phan: Loi + Vo boc loi. Loi lam tu polypropylene/polyester, tiet dien hinh chu nhat, nhieu ranh tac dung thoat nuoc. Vo boc lam bang polyester mong (mang) co tac dung ngan ban va tang them kha nang thoat nuoc cho loi bac tham. Bao gom hai loai: Bac tham dung va bac tham ngang.",
     "1. Gia thanh hop ly voi chat luong quoc te\n2. Nguon nguyen lieu chat luong cao, nhap khau tu nuoc ngoai\n3. Kha nang chong mai mon cuc tot\n4. De dang va rut ngan thoi gian thi cong\n5. Khong bi an mon hay bien chat boi axit, kiem",
     "1. Gia co nen dat yeu: Dat toi 95% on dinh dai han trong thoi gian ngan\n2. On dinh nen: Xu ly nen dat yeu cho cao toc, cau, san bay, duong sat, ben cang\n3. Xu ly moi truong: Xu ly nen dat yeu, dat nho o khu vuc chon lap rac; tay rua dat bi o nhiem",
     "1. Gia co nen dat yeu (dat toi do sau >40m)\n2. On dinh nen duong cao toc, duong dan dau cau, duong bay san bay\n3. Xu ly moi truong dat o nhiem\n4. Xu ly nen dat nhao o khu vuc chon lap rac",
     "ASTM D4595 / TCVN 8871"],

    [4, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "Mang chong tham HDPE",
     "Mang chong tham HDPE",
     "HDPE 0.3mm - 3.0mm",
     "Mang chong tham HDPE Thai Chau duoc san xuat tu hat nhua HDPE voi mat do cao (>0.94g/cm3), co do day tu 0.3mm den 3.0mm. Mang HDPE dang duoc su dung nhieu nhat tren thi truong hien nay, duoc san xuat voi cong nghe tien tien, chat luong uy tin va la mot trong nhung lua chon hang dau cho cac cong trinh co yeu cau khat khe nhat.",
     "1. Tinh tro ly va do ben cao\n2. Tiet kiem chi phi van tai so voi vat lieu dat set truyen thong\n3. Tiet kiem khong gian chon lap rac\n4. Chat luong duoc kiem soat dong nhat\n5. Thi cong nhanh chong",
     "1. Hau nhu khong tham (he so thap hon khoang 1 trieu lan so voi dat set dam nen tot)\n2. Mem deo, co tinh dan dai lon, de lap dat dia hinh phuc tap\n3. Kha nang chieu keo va khang xuyen thung tot\n4. Khong bi xam thuc hoa chat, sinh vat, tr voi acid, kiem, dau\n5. Khong anh huong chat luong nuoc, khong gay tac hai moi truong",
     "1. Bai xu ly rac: Lot day va phu dong bai chon lap rac hop ve sinh\n2. Lot day va mai ho nuoi thuy san: Tao moi truong nuoi kiem soat, chong xi mon\n3. Mang phu noi biogas: Phu ho chat thai chan nuoi, xu ly nuoc thai\n4. Lot day ho chua nuoc khu cong nghiep, ruong muoi\n5. Chong tham de, dap thuy dien, canh quan, san golf\n6. Lot day be chua xang dau, nha may xi, phan bon, hoa chat",
     "ASTM D5199 / D6693 / D4833"],

    [5, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "Mang chong tham Bentonite",
     "Mang chong tham Bentonite (GCL)",
     "APT GCL",
     "Mang chong tham Bentonite hay con goi la mang chong tham GCL (Geosynthetic Clay Liner), san xuat theo cong nghe det xuyen kim va duoc tao nen boi mot lop dat set tong hop (Bentonite) kep giua hai lop vai dia ky thuat, co dac tinh truong no cao se tao ra kha nang chong tham, co tac dung chong thap hon tu 10-100 lan so voi lop dat set luyen day 60-90 cm.",
     "1. De dang thi cong trong khu vuc kho khan hiem hoi\n2. Ngan chan viec gay o nhiem moi truong\n3. Ty le he so tham nuoc rat nho\n4. Chat luong duoc kiem soat dong nhat\n5. Tiet kiem chi phi van tai so voi dat set truyen thong",
     "1. Lot day va dong phu bai chon lap rac\n2. Lot day ao, ho, bon be chua chat nhiem bat\n3. Phu dong bai chon lap rac\n4. Lot day khu chua quang mo, bai xi than\n5. Lot day ao ho khu vui choi, san golf\n6. Chong tham de, dap, kenh muong thuy loi\n7. Chong tham cho cong trinh ngam, ham toa nhan",
     "1. Xu ly moi truong: Lot day bai chon lap rac, ao ho chat nhiem\n2. Chong tham cong trinh ngam\n3. Lot day khu vui choi, san golf\n4. Lot day khu chua quang mo, bai xi than",
     "ASTM D5993 / D5890 / D5084"],

    [6, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "Luoi dia ky thuat",
     "Luoi dia ky thuat APT",
     "PP / PET / Thuy tinh",
     "Luoi dia ky thuat hay con goi la luoi dia la san pham polimer duoc tao thanh tu cac nguyen lieu polimer nhu Polypropylene (PP), high density polyethylene (HDPE) va Polyester (PET). Cac san pham luoi dia ky thuat co cuong do chiu keo cao, bien dang thap la giai phap ly tuong cho cac ket cau dat co cot nhu tuong chan, mai doc dung, gia cuong nen dat yeu co yeu cau tuoi tho cao tren 100 nam.",
     "1. Tinh cai chat voi vat lieu xung quanh, tao lop mong chac chan\n2. Da nang: thich hop voi moi loai dat, da\n3. Tiet kiem khoi luong dao dap\n4. Thoi gian thi cong nhanh\n5. Ben voi tac dong ly hoa moi truong",
     "1. Tuong chan trong luc: Xay tuong chan cao toi 17m voi mai doc den 90 do\n2. Mai doc: Tang kha nang on dinh, khong chuc sut truot (cao toi 50m)\n3. Duong dan dau cau: Tang kha nang chiu tai, tiet kiem khong gian\n4. Lien ket coc: Tao gan do, truyen tai trong hieu qua\n5. Tao luoi do tren nen nhieu hoc trong",
     "1. Xay dung tuong chan trong luc\n2. On dinh mai doc (cao toi 50m)\n3. Duong dan dau cau\n4. Lien ket coc gia co\n5. Phu nen co nhieu hoc trong\n6. Tang ma sat tren mai doc",
     "ASTM D4595 / TCVN 8485"],

    [7, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "O dia ky thuat (Geocell)",
     "O dia ky thuat Geocell",
     "Geocell",
     "O dia ky thuat duoc tao thanh tu nhung tam HDPE lien ke va lien tuc duoc han nhiet voi nhau, nhung o nay sau khi duoc do dat/da/soi se tao thanh mot ket cau co kha nang gia cuong nen dat, mai doc de chong xi mon.",
     "1. Cau truc to ong lien hop chac chan\n2. Gia cuong nen dat hieu qua\n3. Chong xi mon mai doc\n4. Linh hoat trong thi cong",
     "1. Gia cuong nen dat\n2. On dinh mai doc\n3. Chong xi mon be mat",
     "1. Gia cuong mai doc cao\n2. On dinh nen duong, cau\n3. Chong xi mon kenh muong, be bo",
     "ASTM D4595"],

    [8, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "Ong dia ky thuat (Geotube)",
     "Ong dia ky thuat Geotube",
     "Geotube",
     "Ong dia ky thuat duoc lam tu vai dia ky thuat det cuong luc cao Polyester hoac Polypropylene, co kha nang khang chiu duoc moi truong nuoc bien, khang tia cuc tlim (UV) va chiu duoc do pH.",
     "1. Chiu moi truong nuoc bien\n2. Khang tia cuc tlim\n3. Cau truc mem mai theo dia hinh\n4. De thiet ke va thi cong",
     "1. De pha song ngoai khai\n2. Ke, bao bao, loi de\n3. Xay dung ha tang hang hai\n4. Ung pho lu lut",
     "1. Lam de pha song ngoai khai\n2. Lam ke, bao bao\n3. Lam loi de, tuong chan\n4. Xay dung cong trinh ha tang, hang hai\n5. Ung pho su co lu lut",
     "ASTM D4595 / TCVN 8485"],

    [9, "Cong ty TNHH XNK Thai Chau", "thaichau.vn",
     "Vai khong det PET Spunbond",
     "Vai khong det PET Spunbond",
     "PET Spunbond",
     "Thai Chau la don vi dau tien cua Viet Nam san xuat vai khong det PET Spunbond. Duoc san xuat bang chip polyester nguyen sinh thong qua cac quy trinh keo soi nong chay, luc keo toc do cao va ket tinh lam mat. Dac trung boi do ben cao, kha nang chiu nhiet do cao (tren 180 do C), chong lao hoa va on dinh.",
     "1. Do ben keo cao\n2. Chiu duoc nhiet do cao (tren 180 do C)\n3. Chong rac va dam thung cao\n4. Than thien voi moi truong\n5. La don vi dau tien VN san xuat",
     "1. Y te: Vai phau thuat, khan trai giuong, mu dung mot lan\n2. Dong goi: Tui mua sam, khan trai ban, giay dan tuong\n3. Gia dung: Vo goi, vo nem, sofa\n4. Cong nghiep: Bao ve cay trong khoi con trung, thoi tiet lanh\n5. Su dung trong dia ky thuat",
     "1. San pham y te, suc khoe\n2. Dong goi, van phong pham\n3. Noi that gia dung\n4. Nong nghiep, cong nghiep\n5. Dia ky thuat",
     "Tieu chuan noi bo"],

    # ═══════════════════════════════════════════════════════
    # BACTHAM (HUNG PHU) - bactham.vn
    # ═══════════════════════════════════════════════════════
    [10, "Cong ty CP Vat tu Cong trinh Hung Phu", "bactham.vn",
     "Vai dia ky thuat khong det",
     "Vai dia ky thuat khong det ART",
     "ART 7 - ART 28",
     "Vai dia ky thuat khong det ART duoc san xuat tu soi polypropylene chat luong cao bang phuong phap xuyen kim, can nhiet. Hao thanh Carbon den 2-3%. Kho rong 4m. Da dang cuong luc tu 7kN/m den 28kN/m, phu hop voi moi yeu cau ky thuat.",
     "1. Da dang cuong luc 7-28kN/m\n2. Chiu keo, xuyen thung cao\n3. Chong mai mon tot\n4. Gia thanh hop ly\n5. Hang phan phoi uy tin",
     "1. Phan cach: Ngan chan mat vat lieu giua cac lop\n2. Loc: Tranh xi mon tu vat lieu hat min\n3. Tieu thoat: Thu tieu tan nuoc nhanh\n4. Bao ve: Ket hop voi ro da, be tong\n5. Gia cuong: Truyen cho dat cuong do chiu keo",
     "1. Gia cuong nen dat yeu\n2. Loc va tieu thoat nuoc\n3. Phan cach lop vat lieu\n4. Bao ve mai doc, de, dap\n5. Xu ly moi truong",
     "ASTM D4595 / TCVN 8485"],

    [11, "Cong ty CP Vat tu Cong trinh Hung Phu", "bactham.vn",
     "Vai dia ky thuat det",
     "Vai dia ky thuat det GET",
     "GET 10 - GET 300",
     "Vai dia ky thuat det cuong luc cao GET duoc san xuat tu soi polypropylene hoac polyester, mang lai hieu suat toi uu khi ung dung trong cac du an on dinh va gia cuong nen dat yeu. Cuong luc chiu keo va suat dan hoi cao, kha nang loc va thoat nuoc tot.",
     "1. Cuong luc chiu keo cao (100-300 kN/m)\n2. Suat dan hoi cao\n3. Kha nang loc va thoat nuoc tot\n4. Khang UV >70%\n5. Do gian khi dut <=15%",
     "1. Gia cuong nen dat yeu: Tranh bien dang nen dat\n2. Loc va thoat nuoc: He so tham cao\n3. Chong xi mon: Giu dat khong bi xi mon\n4. Phan cach: Phan tach cac lop vat lieu",
     "1. Gia cuong nen duong bo, duong sat\n2. Xay dung ke song, bien\n3. Loc va thoat nuoc\n4. Xu ly moi truong\n5. Xay dung bai rac, bai dung cu",
     "ASTM D4595 / TCVN 8485"],

    [12, "Cong ty CP Vat tu Cong trinh Hung Phu", "bactham.vn",
     "Ro da - Tham da",
     "Ro da (Gabion)",
     "Gabion 8x10 / 6x8 / 10x12",
     "Ro da hay con goi la Gabion, la cau truc duoc tao thanh tu cac lung luoi thep dan xen nhau, ben trong chua day da hoac cac vat lieu tu nhien khac. Lung luoi thuong lam tu thep ma kem hoac thep boc nhua PVC de chong an mon. Mat luoi: 8x10cm, 6x8cm, 10x12cm. Day dan ma kem nhe theo TCVN 2053-1993, day dan ma kem boc PVC theo TCVN 10335-2014.",
     "1. Dan tren may chuyen dung\n2. Tieu chuan mat luoi xoan kep\n3. Nha xuong 3.000 m2\n4. Cong suat 120 Tan/thang (5 may dan)\n5. Day ma kem hoac boc PVC",
     "1. Gia co nen mong va bao ve mai doc\n2. Xay dung be ke va cong trinh thuy loi\n3. Bao ve moi truong\n4. Tao canh quan tu nhien",
     "1. Tuong chan dat, ke be\n2. Be ke song, kenh, ho\n3. Chong xi mon be mat\n4. Tao canh quan vuon hoa\n5. Xu ly moi truong",
     "TCVN 2053-1993 / TCVN 10335-2014"],

    [13, "Cong ty CP Vat tu Cong trinh Hung Phu", "bactham.vn",
     "Mang chong tham HDPE",
     "Mang chong tham HDPE HSE",
     "HSE 0.3mm - 2.0mm",
     "Mang chong tham HDPE cua Hung Phu la hang phan phoi cap 1 cua nha may Aritex. Thuong hieu nhan dang la HSE. Kho hang tai Long An. HDPE (High density Polyethylene) la nhua PE co trong luong phan tu lon chiem 97.5%, 2.5% than hoat tinh va cac hoat chat chong oxi hoa. Kho muoi hon 20 nam (san xuat trong nuoc) va den 50 nam (nhap khau).",
     "1. Dai ly cap 1 cua Aritex\n2. Kho hang tai Long An\n3. Thi cong lap dat chuyen nghiep\n4. May han Comet Leister (Thuy Si)\n5. Giao hang dung tien do",
     "1. Lot day ao, ho\n2. Chong tham biogas\n3. Chon lap rac thai\n4. Chong tham ho chua nuoc thai\n5. Lot be chua xang dau",
     "1. Lot ho tom nuoi thuy san\n2. Ho biogas xu ly mo truong\n3. Bao che bai chon lap rac\n4. Chong tham ho chua nuoc\n5. Lot be chua hoa chat, xang dau\n6. Xu ly nuoc thai cong nghiep",
     "ASTM D5199 / D6693 / D4833"],

    [14, "Cong ty CP Vat tu Cong trinh Hung Phu", "bactham.vn",
     "Vi nhua thoat nuoc",
     "Vi nhua thoat nuoc / VersiCell",
     "VersiCell",
     "Vi nhua thoat nuoc la giai phap xanh cho do thi tuong lai. La loai san pham xay dung co cau truc o ong, dung de thoat nuoc, tao lop khong khi giua cac lop vat lieu. Pho bien trong cac cong trinh san vuon, mai xanh, san tiec.",
     "1. Giai phap thoat nuoc hieu qua\n2. Than thien moi truong\n3. De dang lap dat\n4. Nhe, de van chuyen",
     "1. Thoat nuoc san, mai\n2. Tao lop khong khi\n3. Ngan ngua tich tuy nuoc",
     "1. San vuon, mai xanh\n2. San tiec, san the thao\n3. Cong trinh dan cu",
     "Tieu chuan noi bo"],

    [15, "Cong ty CP Vat tu Cong trinh Hung Phu", "bactham.vn",
     "Tui vai trong Cay",
     "Tui vai trong Cay",
     "T80 / T20-T60",
     "Tui vai trong co quai xach, dung de trong cay, giong cay. Kich thuoc da dang tu T20 den T80. Combo 20 cai. Phu hop trong nuoi trong va lam vuon.",
     "1. Co quai xach tien loi\n2. Da dang kich thuoc\n3. Gia re, tiet kiem\n4. De su dung",
     "1. Trong cay, giong cay\n2. Lam vuon\n3. Trang tri",
     "1. Vuon gia dinh\n2. Trang tri san vuon\n3. Nong nghiep",
     "Tieu chuan noi bo"],

    # ═══════════════════════════════════════════════════════
    # ARITEX - aritex.com.vn
    # ═══════════════════════════════════════════════════════
    [16, "Cong ty CP Vai dia ky thuat Viet Nam", "aritex.com.vn",
     "Vai dia ky thuat khong det",
     "Vai dia ky thuat khong det ART",
     "ART 7 - ART 28 (Pho thong)\nART 9D - ART 28D (Loai D)\nART 700G, 900G, 12A",
     "Vai dia ky thuat khong det ART duoc san xau tu xo polypropylene chat luong cao bang phuong phap xuyen kim, can nhiet. Tao nen mot lop vat lieu chac chan, giu duoc su on dinh ve kich thuoc va co do ben cao khi ung dung trong cac du an xay dung. Nha san xuat dau tien va lon nhat Viet Nam.",
     "1. Nha san xuat dau tien & lon nhat Viet Nam\n2. Pho thong: 7-28kN/m, O95: 60-150um\n3. Loai D: 9.5-28kN/m, Day 1.2-3.2mm\n4. Theo thiet ke du an: 700G/900G/12A\n5. Khob rong 4m",
     "1. Phan cach: Phan tach lop vat lieu\n2. Gia cuong: Tang cuong do chiu keo\n3. Bao ve: Chong xam thuc, xi mon\n4. Loc: Ngan chan hat dat bi di chuyen\n5. Tieu thoat: Thu tieu tan nuoc",
     "1. Phan cach, gia cuong, bao ve\n2. Loc va tieu thoat nuoc\n3. Xu ly nen dat yeu\n4. Cong trinh giao thong\n5. Xu ly moi truong",
     "ASTM D4595 / TCVN 8485"],

    [17, "Cong ty CP Vai dia ky thuat Viet Nam", "aritex.com.vn",
     "Vai dia ky thuat det",
     "Vai dia ky thuat det GET",
     "GET 10 - GET 300",
     "Duoc san xau tu cac nguyen lieu chat luong cao, vai dia ky thuat det cuong luc cao GET ket hop cac dac tinh cuong luc chiu keo va suat dan hoi cao, kha nang loc va thoat nuoc tot, co the dap ung yeu cau ky thuat nhieu du an, mang lai hieu suat toi uu khi su dung.",
     "1. Cuong luc chiu keo cao (100-300 kN/m)\n2. Suat dan hoi cao\n3. Kha nang loc thoat nuoc tot\n4. Khang UV >70%\n5. Kho rong 3.5m hoac 5.4m",
     "1. Gia cuong nen dat yeu\n2. Loc va thoat nuoc\n3. Chong xi mon\n4. Phan cach lop vat lieu",
     "1. Gia cuong nen duong bo, duong sat\n2. Xay dung ke song, bien\n3. Loc va thoat nuoc\n4. Xu ly moi truong\n5. Xay dung bai rac",
     "ASTM D4595 / TCVN 8485"],

    [18, "Cong ty CP Vai dia ky thuat Viet Nam", "aritex.com.vn",
     "Mang chong tham HDPE",
     "Mang chong tham HDPE HSE",
     "HSE 0.3 - HSE 2.0",
     "Cong ty CP Vai dia ky thuat Viet Nam la nha san xau mang chong tham HDPE dau tien va lon nhat tai Viet Nam. Mang HDPE HSE duoc san xau tu hat nhua polyethylene nguyen sinh ty trong cao va hat carbon den; co be mat, do day va kich co khac nhau. Do day pho bien tu 0.5mm den 2.0mm. Tuoi tho tren 25 nam.",
     "1. Nha san xau HDPE dau tien & lon nhat Viet Nam\n2. San xau tu hat nhua nguyen sinh\n3. Khang UV, khang hoa chat\n4. Tuoi tho >100 nam (ly thuyet)\n5. Kho rong 8m, cat theo yeu cau",
     "1. Lot ho nuoi thuy san\n2. Chon lap rac thai\n3. Xu ly moi truong o nhiem dioxin\n4. Lam hami biogas\n5. Lot be chua xang dau",
     "1. Lot ho nuoi tom, ca\n2. Bao che bai chon lap rac\n3. Xu ly moi truong\n4. Ho biogas\n5. Lot be chua xang dau, hoa chat\n6. Chong tham de, dap, san golf",
     "ASTM D5199 / D6693 / D4833"],

    [19, "Cong ty CP Vai dia ky thuat Viet Nam", "aritex.com.vn",
     "Bac tham",
     "Bac tham dung & ngang",
     "VID 75, RID 75, RID 4.0, VID 4.5 (dung)\nRID 200, RID 300 (ngang)",
     "Bac tham la vat lieu dia ky thuat dung de thoat nuoc muc dich gia tang qua trinh co ket cua nen mong. Cau tao tu hai lop: lop vo bang vai dia ky thuat khong det polyester, lop loi thoat nuoc san xau tu nhua polypropylene.",
     "1. Loi PP, vo loc polyester\n2. Chong vi khong, khong an mon\n3. Bien dang theo dia hinh\n4. Duy tri kha nang thoat nuoc\n5. Be rong 100-300mm",
     "1. Bac tham dung: Xu ly gia co nen dat yeu\n2. Bac tham ngang: Thoat nuoc ngang\n3. Chong an mon, bien chat\n4. Kha nang tuong tich voi nhieu loai dat",
     "1. Xu ly gia co nen dat yeu\n2. Duong cao toc, san bay, duong sat\n3. Ben cang, kho xang dau\n4. Xu ly moi truong dat o nhiem",
     "ASTM D4595 / D4716 / D4491"],

    [20, "Cong ty CP Vai dia ky thuat Viet Nam", "aritex.com.vn",
     "Mang chong tham set tong hop",
     "Mang chong tham set tong hop ART Bentonite",
     "ART 3000 / ART 4000 / ART 4700",
     "Mang chong tham set tong hop ART Bentonite duoc cau tao voi lop phu be mat bang vai dia ky thuat khong det, lop giua la bentonite tu nhien va lop lot day la vai dia ky thuat, khi gap nuoc se truong no tao ra mang keo co tac dung chong tham hieu qua.",
     "1. Cau tao 3 lop: vai + Bentonite + vai\n2. Truong no cao khi gap nuoc\n3. He so tham cuc thap (<=5x10-11 m/s)\n4. Chiu khac boc >=65N\n5. Chi so truong no >24ml/2g",
     "1. Chon lap chat thai\n2. Niem phong kinh phong xa, chat doc\n3. Lot ao ho chat nhiem\n4. Xu ly moi truong",
     "1. Xu ly moi truong: Chon lap rac, chat thai\n2. Xu ly dat o nhiem\n3. Lot ao ho, bon be chua\n4. Chong tham cong trinh ngam",
     "ASTM D5084 / D5993 / D5890"],

    [21, "Cong ty CP Vai dia ky thuat Viet Nam", "aritex.com.vn",
     "Ong dia ky thuat",
     "Ong / Tui dia ky thuat",
     "Ong 50x11.2m\nTui 10x4.4m",
     "Ong dia ky thuat Aritex duoc lam tu vai dia ky thuat det cuong luc cao polyester hoac polypropylene, co kha nang khang chiu duoc moi truong nuoc bien, khang tia cuc tlim (UV) va chiu duoc do pH. Thiet ke rieng biet va vai det co cau truc dac biet, co kha nang thoat nuoc tot.",
     "1. Lam tu vai det cuong luc cao\n2. Khang nuoc bien, UV, chiu pH\n3. Thiet ke theo du an\n4. ISO 9001\n5. De thiet ke va thi cong",
     "1. De pha song ngoai khai\n2. Ke, bao bao, loi de\n3. Tuong chan\n4. Xay dung ha tang hang hai\n5. Ung pho lu lut",
     "1. De pha song ngoai khai\n2. Lam ke, bao bao\n3. Lam loi de, tuong chan\n4. Xay dung cong trinh ha tang\n5. Ung pho su co lu lut",
     "ASTM D4595 / TCVN 8485"],

    [22, "Cong ty CP Vai dia ky thuat Viet Nam", "aritex.com.vn",
     "Luoi dia ky thuat",
     "Luoi dia ky thuat",
     "Don truc / Hai truc / Cot soi",
     "Luoi dia ky thuat la vat lieu tong hop dung trong xay dung ha tang, co chuc nang gia co nen dat, tang cuong on dinh cong trinh, kiem soat xi mon va thoat nuoc.",
     "1. Don truc: Chiu luc 1 chieu\n2. Hai truc: Chiu luc 2 chieu\n3. Cot soi: Ket hop vat lieu khac\n4. Gia co nen dat hieu qua",
     "1. Gia co mai doc\n2. On dinh nen mong\n3. Kiem soat xi mon\n4. Tang do lien ket lop dat",
     "1. Gia co mai doc\n2. On dinh nen mong, nen duong\n3. Kiem soat xi mon be mat\n4. Tang do on dinh cong trinh",
     "ASTM D4595"],

    [23, "Cong ty CP Vai dia ky thuat Viet Nam", "aritex.com.vn",
     "O dia ky thuat",
     "O dia ky thuat Acel (Geocell)",
     "Acel",
     "O dia ky thuat Acel la giai phap tien tinh dang duoc ap dung tren khap toan cau. La ket cau dia ky thuat lien hop vung chac hinh to ong, duoc tao thanh cach han nhiet cac tam HDPE, o dia ky thuat sau khi duoc do cac vat lieu chen (cat, dat, soi,...) se tao thanh mot tam co chuc nang gia co mai doc hay bao ve duong, bai.",
     "1. Cau truc to ong lien hop\n2. Han nhiet tam HDPE\n3. Gia cuong chiu tai\n4. Chong xi mon be mat\n5. Carbon den >=1.5%, Ti trong >=0.94",
     "1. Gia cuong nen dat\n2. Bao ve be mat\n3. Chong xi mon\n4. Giam toc do dong chay mat",
     "1. Gia co mai doc\n2. On dinh nen mong\n3. Chong xi mon kenh muong\n4. Bao ve nen duong, bai",
     "ASTM D1603 / D1505 / D5199"],
]

write_data(ws1, products)
ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}{len(products)+1}"

# ═══════════════════════════════════════════════════════
# SHEET 2: THONG SO KY THUAT CHI TIET
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Thong so ky thuat")

h2 = ["Nha cung cap", "Nhom san pham", "Ma SP", "Cuong do chiu keo (kN/m)", "Do dan dai (%)",
      "CBR / Khang thung (N)", "He so tham", "Kich thuoc lo O95", "Kho rong (m)",
      "Do day (mm)", "Trong luong (g/m2)", "Tieu chuan"]
w2 = [18, 25, 18, 22, 14, 20, 18, 16, 12, 12, 16, 22]
setup_sheet(ws2, h2, w2)

specs = [
    # ARITEX - Vai khong det pho thong
    ["Aritex", "Vai KDT khong det", "ART 7", "7.0", "40-75", "1200", "210 L/m2/s", "150 um", "4", "", "100-120", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 9", "9.0", "40-75", "1500", "170 L/m2/s", "120 um", "4", "", "130-150", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 11", "11.0", "40-75", "1700", "150 L/m2/s", "115 um", "4", "", "150-170", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 12", "12.0", "40-75", "1900", "140 L/m2/s", "110 um", "4", "", "170-190", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 14", "14.0", "40-75", "2100", "125 L/m2/s", "100 um", "4", "", "200-220", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 15", "15.0", "40-75", "2400", "120 L/m2/s", "90 um", "4", "", "220-240", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 17", "17.0", ">50", "2700", "90 L/m2/s", "80 um", "4", "", "250-280", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 20", "20.0", ">50", "2900", "80 L/m2/s", "75 um", "4", "", "300-340", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 22", "22.0", ">50", "3200", "75 L/m2/s", "75 um", "4", "", "340-380", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 24", "24.0", ">50", "3800", "70 L/m2/s", "70 um", "4", "", "380-420", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 25", "25.0", ">50", "4000", "60 L/m2/s", "70 um", "4", "", "400-440", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT khong det", "ART 28", "28.0", ">50", "4500", "50 L/m2/s", "60 um", "4", "", "440-480", "ASTM D4595 / TCVN 8485"],
    # ARITEX - Vai khong det loai D
    ["Aritex", "Vai KDT khong det (D)", "ART 9D", "9.5", ">40", "1500", "30x10-4 m/s", "180 um", "", "1.2", "150-170", "ASTM D4595"],
    ["Aritex", "Vai KDT khong det (D)", "ART 11D", "11.5", ">40", "1800", "30x10-4 m/s", "150 um", "", "1.5", "180-200", "ASTM D4595"],
    ["Aritex", "Vai KDT khong det (D)", "ART 15D", "15.0", ">40", "2400", "30x10-4 m/s", "110 um", "", "1.9", "250-280", "ASTM D4595"],
    ["Aritex", "Vai KDT khong det (D)", "ART 22D", "22.0", ">40", "3300", "30x10-4 m/s", "85 um", "", "2.5", "380-420", "ASTM D4595"],
    ["Aritex", "Vai KDT khong det (D)", "ART 28D", "28.0", ">40", "4500", "30x10-4 m/s", "60 um", "", "3.2", "440-480", "ASTM D4595"],
    # ARITEX - Vai det
    ["Aritex", "Vai KDT det", "GET 10", ">=100/50", "<=15", ">=4500", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT det", "GET 15", ">=150/50", "<=15", ">=5500", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT det", "GET 20", ">=200/50", "<=15", ">=7000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT det", "GET 40", ">=400/50", "<=15", ">=12000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT det", "GET 100", ">=100/100", "<=15", ">=6000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT det", "GET 200", ">=200/200", "<=15", ">=15000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vai KDT det", "GET 300", ">=300/300", "<=15", ">=18000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "", "ASTM D4595 / TCVN 8485"],
    # ARITEX - HDPE
    ["Aritex", "Mang HDPE", "HSE 0.3", "8", "600", "105 (xuyen thung)", "", "", "8", "0.3", "", "ASTM D5199 / D6693"],
    ["Aritex", "Mang HDPE", "HSE 0.5", "14", "700", "176 (xuyen thung)", "", "", "8", "0.5", "", "ASTM D5199 / D6693"],
    ["Aritex", "Mang HDPE", "HSE 0.75", "21", "700", "264 (xuyen thung)", "", "", "8", "0.75", "", "ASTM D5199 / D6693"],
    ["Aritex", "Mang HDPE", "HSE 1.0", "28", "700", "352 (xuyen thung)", "", "", "8", "1.0", "", "ASTM D5199 / D6693"],
    ["Aritex", "Mang HDPE", "HSE 1.5", "43", "700", "540 (xuyen thung)", "", "", "8", "1.5", "", "ASTM D5199 / D6693"],
    ["Aritex", "Mang HDPE", "HSE 2.0", "57", "700", "705 (xuyen thung)", "", "", "7", "2.0", "", "ASTM D5199 / D6693"],
    # ARITEX - Bentonite
    ["Aritex", "Mang set tong hop", "ART 3000", "", "", "", "<=5x10-11 m/s", "", "", "", "2700+180+110", "ASTM D5084"],
    ["Aritex", "Mang set tong hop", "ART 4000", "", "", "", "<=3x10-11 m/s", "", "", "", "3700+180+110", "ASTM D5084"],
    ["Aritex", "Mang set tong hop", "ART 4700", "", "", "", "<=5x10-11 m/s", "", "", "", "4700+180+110", "ASTM D5084"],
    # ARITEX - Bac tham
    ["Aritex", "Bac tham dung", "VID 75", ">1.6", ">20", "", ">=1.4x10-4 m/s", "<0.075 mm", "100mm", "", "", "ASTM D4595 / D4491"],
    ["Aritex", "Bac tham dung", "RID 75", ">1.7", ">20", "", ">=1.4x10-4 m/s", "<0.075 mm", "100mm", "", "", "ASTM D4595 / D4491"],
    ["Aritex", "Bac tham dung", "RID 4.0", ">1.9", ">20", "", ">=1.4x10-4 m/s", "<0.075 mm", "100mm", "", "", "ASTM D4595 / D4491"],
    ["Aritex", "Bac tham ngang", "RID 200", "", "<25", ">250 (nen)", ">1.4x10-4 m/s", "<0.075 mm", "200mm", "8.0", "", "ASTM D4595 / D4491"],
    ["Aritex", "Bac tham ngang", "RID 300", "", "<25", ">250 (nen)", ">1.4x10-4 m/s", "<0.075 mm", "300mm", "8.0", "", "ASTM D4595 / D4491"],
    # ARITEX - Geocell
    ["Aritex", "O dia ky thuat", "Acel", "", "", ">=1000 (moi han)", "", "", "", "1.5", ">=1.41", "ASTM D1603 / D1505"],
    # HUNG PHU
    ["Hung Phu", "Vai KDT khong det", "ART 7", "7", "40-75", "1200", "", "150 um", "4", "", "", ""],
    ["Hung Phu", "Vai KDT khong det", "ART 9", "9", "40-75", "1500", "", "120 um", "4", "", "", ""],
    ["Hung Phu", "Vai KDT khong det", "ART 12", "12", "40-75", "1900", "", "110 um", "4", "", "", ""],
    ["Hung Phu", "Vai KDT khong det", "ART 15", "15", "40-75", "2400", "", "90 um", "4", "", "", ""],
    ["Hung Phu", "Vai KDT khong det", "ART 25", "25", ">50", "4000", "", "70 um", "4", "", "", ""],
    ["Hung Phu", "Vai KDT det", "GET 10", "100/50", "<=15", ">=4500", "", "", "", "", "", ""],
    ["Hung Phu", "Vai KDT det", "GET 15", "150/50", "<=15", ">=5500", "", "", "", "", "", ""],
    ["Hung Phu", "Vai KDT det", "GET 20", "200/50", "<=15", ">=7000", "", "", "", "", "", ""],
    ["Hung Phu", "Vai KDT det", "GET 100", "100/100", "<=15", ">=6000", "", "", "", "", "", ""],
    ["Hung Phu", "Vai KDT det", "GET 200", "200/200", "<=15", ">=15000", "", "", "", "", "", ""],
    ["Hung Phu", "Vai KDT det", "GET 300", "300/300", "<=15", ">=18000", "", "", "", "", "", ""],
    ["Hung Phu", "Mang HDPE", "HSE 0.5", "14", "700", "176", "", "", "8", "0.5", "", "ASTM"],
    ["Hung Phu", "Mang HDPE", "HSE 0.75", "21", "700", "264", "", "", "8", "0.75", "", "ASTM"],
    ["Hung Phu", "Mang HDPE", "HSE 1.0", "28", "700", "352", "", "", "8", "1.0", "", "ASTM"],
    ["Hung Phu", "Mang HDPE", "HSE 1.5", "43", "700", "540", "", "", "8", "1.5", "", "ASTM"],
    ["Hung Phu", "Mang HDPE", "HSE 2.0", "57", "700", "705", "", "", "7", "2.0", "", "ASTM"],
    ["Hung Phu", "Ro da", "Gabion 8x10", "", "", "", "", "Mat luoi 8x10cm", "", "", "", "TCVN 2053"],
    ["Hung Phu", "Ro da", "Gabion 6x8", "", "", "", "", "Mat luoi 6x8cm", "", "", "", "TCVN 2053"],
    ["Hung Phu", "Ro da", "Gabion 10x12", "", "", "", "", "Mat luoi 10x12cm", "", "", "", "TCVN 2053"],
]

write_data(ws2, specs)
ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{len(specs)+1}"

# ═══════════════════════════════════════════════════════
# SHEET 3: SO SANH NHA CUNG CAP
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("So sanh nha cung cap")

h3 = ["Nhom san pham", "Thai Chau (thaichau.vn)", "Hung Phu (bactham.vn)", "Aritex (aritex.com.vn)"]
w3 = [28, 40, 40, 40]
setup_sheet(ws3, h3, w3)

comp = [
    ["Vai KDT khong det", "APT 7-80kN/m\n100% PP, xuyen kim\nCong nghe Han Quoc\nNha san xuat", "ART 7-28kN/m\nPho thong & Loai D\nHang phan phoi", "ART 7-28kN/m\nPho thong, Loai D, Theo thiet ke\nNha san xau dau tien & lon nhat VN"],
    ["Vai KDT det", "DML\n100% Polyester\nDay chuyen tu dong\nNha san xuat", "GET 10-300 kN/m\nPP/PET\nHang phan phoi", "GET 10-300 kN/m\nPP/PET\nNha san xuat"],
    ["Mang chong tham HDPE", "0.3-3.0mm\n>0.94g/cm3\nNha san xuat", "HSE 0.3-2.0mm\nThuong hieu Aritex\nDai ly cap 1, kho Long An", "HSE 0.3-2.0mm\nNha san xau dau tien & lon nhat VN\nKho rong 8m"],
    ["Bac tham", "VID/RID\nDung & Ngang\nNha san xuat\nDo sau >40m", "Bac tham dung & ngang\nCung cap dich vu", "VID 75, RID 75, RID 4.0\nVID 4.5, RID 200, RID 300\nNha san xuat"],
    ["Mang chong tham set (GCL)", "APT GCL\nXuyen kim\nNha san xuat\nTham 10-100 lan dat set", "Khong co thong tin", "ART 3000/4000/4700\n3 lop: vai + Bentonite + vai\nNha san xuat"],
    ["Luoi dia ky thuat", "PP, PET, Thuy tinh\nDon truc, da truc\nNha san xuat", "Khong co thong tin", "Don truc, hai truc, cot soi\nGia co nen, mai doc"],
    ["O dia ky thuat (Geocell)", "HDPE han nhiet\nGia co mai doc\nNha san xuat", "Khong co thong tin", "Acel - HDPE\nCarbon den >=1.5%\nTi trong >=0.94"],
    ["Ro da / Tham da", "Khong co thong tin", "Gabion\nMat luoi 6-12cm\nMa kem / PVC\nNha san xuat (120 Tan/thang)", "Khong co thong tin"],
    ["Ong dia ky thuat (Geotube)", "PET/PP cuong luc cao\nKhang nuoc bien, UV\nNha san xuat", "Khong co thong tin", "50x11.2m (ong)\n10x4.4m (tui)\nISO 9001"],
    ["Vi nhua thoat nuoc", "Khong co thong tin", "VersiCell\nGiai phap thoat nuoc\nSan vuon, mai xanh", "Khong co thong tin"],
    ["Tui vai trong cay", "Khong co thong tin", "T80 / T20-T60\nCo quai xach\nCombo 20 cai", "Khong co thong tin"],
    ["Vai khong det PET Spunbond", "PET Spunbond\n100% polyester nguyen sinh\nDau tien VN san xuat\nChiu nhiet >180 do C", "Khong co thong tin", "Khong co thong tin"],
]

write_data(ws3, comp)

# ═══════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════
output = "san_pham_dia_ky_thuat_chi_tiet.xlsx"
wb.save(output)
print(f"Done: {output}")
print(f"  Sheet 1 - Danh muc san pham: {len(products)} san pham")
print(f"  Sheet 2 - Thong so ky thuat: {len(specs)} dong")
print(f"  Sheet 3 - So sanh nha cung cap: {len(comp)} nhom")
