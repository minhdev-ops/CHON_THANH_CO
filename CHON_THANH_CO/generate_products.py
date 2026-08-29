#!/usr/bin/env python3
"""
Scrape and compile product data from 3 geosynthetic websites into Excel.
Sites: thaichau.vn, bactham.vn, aritex.com.vn
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
company_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
company_font = Font(name='Arial', bold=True, size=11, color='2F5496')
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ═══════════════════════════════════════════════════════
# SHEET 1: TỔNG HỢP SẢN PHẨM
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Tổng hợp sản phẩm"

headers = ["STT", "Nhà cung cấp", "Website", "Nhóm sản phẩm", "Tên sản phẩm", "Mã sản phẩm", "Thông số kỹ thuật", "Đặc điểm nổi bật", "Ứng dụng"]
col_widths = [6, 22, 22, 25, 35, 20, 55, 45, 45]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── DATA ──
products = [
    # ═══ THÁI CHÂU (thaichau.vn) ═══
    # 1. Vải địa kỹ thuật không dệt APT
    {
        "company": "Công ty TNHH XNK Thái Châu",
        "website": "thaichau.vn",
        "category": "Vải địa kỹ thuật không dệt",
        "name": "Vải địa kỹ thuật không dệt APT",
        "code": "APT 7 – APT 80",
        "specs": "Cường độ chịu kéo: 7kN – 80kN/m (100g/m2 – 1200g/m2)\nSản xuất từ 100% PP, xuyên kim, phụ gia UV\nDây chuyền tự động, công nghệ Hàn Quốc",
        "features": "Đa dạng cường lực, tính ổn định cao\nGiá hợp lý với chất lượng quốc tế\nNguyên liệu PP nhập khẩu\nHỗ trợ kỹ thuật chuyên môn",
        "applications": "Phân cách, lọc, tiêu thoát nước\nBảo vệ, gia cường nền đất yếu\nGia cố cầu đường, sân bay\nBảo vệ màng chống thấm"
    },
    # 2. Vải địa kỹ thuật dệt DML
    {
        "company": "Công ty TNHH XNK Thái Châu",
        "website": "thaichau.vn",
        "category": "Vải địa kỹ thuật dệt",
        "name": "Vải địa kỹ thuật dệt DML",
        "code": "DML",
        "specs": "Sản xuất tại nhà máy APT\nDây chuyền tự động, tiêu chuẩn Hàn Quốc\n100% nguyên liệu Polyester",
        "features": "Cường lực chịu kéo cao\nKhả năng lọc và thoát nước tốt\nĐộ bền môi trường cao",
        "applications": "Gia cố nền đất yếu\nLọc và thoát nước\nChống xói mòn kè sông, biển"
    },
    # 3. Bấc thấm APT
    {
        "company": "Công ty TNHH XNK Thái Châu",
        "website": "thaichau.vn",
        "category": "Bấc thấm",
        "name": "Bấc thấm PVD APT",
        "code": "VID / RID",
        "specs": "Bấc thấm đứng và bấc thấm ngang\nCấu tạo: Lõi PP/PET + Vỏ polyester\nCó thể đóng xuống độ sâu >40m",
        "features": "Nguyên liệu nhập khẩu chất lượng cao\nChống mài mòn cực tốt\nGiá hợp lý chất lượng quốc tế\nRút ngắn thời gian thi công",
        "applications": "Gia cố nền đất yếu\nỔn định nền cầu đường, sân bay\nXử lý môi trường đất ô nhiễm"
    },
    # 4. Màng chống thấm Bentonite (GCL)
    {
        "company": "Công ty TNHH XNK Thái Châu",
        "website": "thaichau.vn",
        "category": "Màng chống thấm Bentonite",
        "name": "Màng chống thấm Bentonite GCL",
        "code": "APT GCL",
        "specs": "Công nghệ dệt xuyên kim\nLớp Bentonite kẹp giữa 2 lớp vải địa kỹ thuật\nHệ số thấm thấp hơn 10-100 lần so với đất sét dày 60-90cm",
        "features": "Dễ thi công vùng khó khăn\nNgăn ô nhiễm môi trường\nChất lượng kiểm soát đồng nhất\nTiết kiệm vận tải",
        "applications": "Lót đáy bãi chôn lấp rác\nLót ao hồ chứa chất nhiễm bẩn\nChống thấm đê đập, kênh mương\nLót đáy khu vui chơi, sân golf"
    },
    # 5. Lưới địa kỹ thuật APT
    {
        "company": "Công ty TNHH XNK Thái Châu",
        "website": "thaichau.vn",
        "category": "Lưới địa kỹ thuật",
        "name": "Lưới địa kỹ thuật APT",
        "code": "PP / PET / Thủy tinh",
        "specs": "Lưới PP, lưới PET, lưới cốt sợi thủy tinh\nLưới ô hình chữ nhật, lưới tam giác\nĐa trục: đơn trục, đa trục",
        "features": "Đa dạng sản phẩm lưới\nCường độ chịu kéo cao, biến dạng thấp\nBền với tác động lý hóa môi trường\nTuổi thọ >100 năm",
        "applications": "Tường chắn trọng lực (cao tới 17m)\nMái dốc (cao tới 50m)\nĐường dẫn đầu cầu\nCải tạo nâng cấp đường cao tốc, sân bay"
    },
    # 6. Ô địa kỹ thuật – Geocell
    {
        "company": "Công ty TNHH XNK Thái Châu",
        "website": "thaichau.vn",
        "category": "Ô địa kỹ thuật (Geocell)",
        "name": "Ô địa kỹ thuật Geocell APT",
        "code": "Geocell",
        "specs": "Tấm HDPE liền kề hàn nhiệt\nSau khi đổ đất/đá/sỏi tạo kết cấu gia cường",
        "features": "Gia cường nền đất\nChống xói mòn mái dốc\nCấu trúc tổ ong linh hoạt",
        "applications": "Gia cố mái dốc\nỔn định nền đường\nChống xói mòn bề mặt"
    },
    # 7. Màng chống thấm HDPE
    {
        "company": "Công ty TNHH XNK Thái Châu",
        "website": "thaichau.vn",
        "category": "Màng chống thấm HDPE",
        "name": "Màng chống thấm HDPE",
        "code": "HDPE 0.3 – 3.0mm",
        "specs": "Từ hạt nhựa HDPE mật độ cao (>0.94g/cm3)\nĐộ dày: 0.3mm – 3.0mm\nCó phụ gia chống lão hóa, UV, kháng hóa chất\nTuổi thọ thiết kế >50 năm",
        "features": "Tính trơ lỳ, độ bền cao\nTiết kiệm chi phí vận tải\nChất lượng kiểm soát đồng nhất\nKhông gây độc hại môi trường",
        "applications": "Bãi xử lý rác\nLót hồ nuôi thủy sản\nMàng phủ hồ biogas\nChống thấm đê đập, sân golf\nLót bể chứa xăng dầu"
    },
    # 8. Ống địa kỹ thuật – Geotube
    {
        "company": "Công ty TNHH XNK Thái Châu",
        "website": "thaichau.vn",
        "category": "Ống địa kỹ thuật (Geotube)",
        "name": "Ống địa kỹ thuật Geotube",
        "code": "Geotube",
        "specs": "Vải địa kỹ thuật dệt cường lực cao\nPolyester hoặc Polypropylene\nKháng nước biển, UV, chịu pH",
        "features": "Chịu môi trường nước biển\nKháng tia cực tím\nCấu trúc mềm mại theo địa hình",
        "applications": "Đê phá sóng ngoài khơi\nKè, bờ bao, lõi đê\nXây dựng hạ tầng hàng hải\nỨng phó lũ lụt"
    },

    # ═══ BẤC THẨM (bactham.vn) ═══
    # 9. Vải địa kỹ thuật không dệt ART
    {
        "company": "Công ty CP Vật tư Công trình Hưng Phú",
        "website": "bactham.vn",
        "category": "Vải địa kỹ thuật không dệt",
        "name": "Vải địa kỹ thuật không dệt ART",
        "code": "ART 7 – ART 28",
        "specs": "ART 7: 7kN/m, CBR 1200N, O95 150μm\nART 9: 9kN/m, CBR 1500N, O95 120μm\nART 12: 12kN/m, CBR 1900N, O95 110μm\nART 15: 15kN/m, CBR 2400N, O95 90μm\nART 25: 25kN/m, CBR 4000N, O95 70μm\nKhổ rộng: 4m",
        "features": "Đa dạng cường lực 7-28kN/m\nChịu kéo, xuyên thủng cao\nChống mài mòn tốt\nGiá thành hợp lý",
        "applications": "Gia cố nền đất yếu\nLọc và tiêu thoát nước\nPhân cách lớp vật liệu\nBảo vệ mái dốc"
    },
    # 10. Vải địa kỹ thuật dệt GET
    {
        "company": "Công ty CP Vật tư Công trình Hưng Phú",
        "website": "bactham.vn",
        "category": "Vải địa kỹ thuật dệt",
        "name": "Vải địa kỹ thuật dệt GET",
        "code": "GET 10 – GET 300",
        "specs": "GET 10: 100/50 kN/m, CBR ≥4500N\nGET 15: 150/50 kN/m, CBR ≥5500N\nGET 20: 200/50 kN/m, CBR ≥7000N\nGET 100: 100/100 kN/m, CBR ≥6000N\nGET 200: 200/200 kN/m, CBR ≥15000N\nGET 300: 300/300 kN/m, CBR ≥18000N\nĐộ giãn: ≤15%, UV >70%",
        "features": "Cường lực chịu kéo cao\nSuất đàn hồi cao\nKhả năng lọc thoát nước tốt\nKháng UV >70%",
        "applications": "Gia cố nền đất yếu\nLọc và thoát nước\nChống xói mòn kè sông, biển\nXây dựng đường bộ, đường sắt"
    },
    # 11. Rọ đá Hưng Phú
    {
        "company": "Công ty CP Vật tư Công trình Hưng Phú",
        "website": "bactham.vn",
        "category": "Rọ đá – Thảm đá",
        "name": "Rọ đá (Gabion)",
        "code": "Gabion",
        "specs": "Mắt lưới: 8x10cm, 6x8cm, 10x12cm\nDây đan: Mạ kẽm nhẹ (TCVN 2053-1993)\nDây đan: Mạ kẽm bọc PVC (TCVN 10335-2014)\nMạ kẽm nặng: 220-280 g/m2\nCông suất: 120 Tấn/tháng",
        "features": "Đan trên máy chuyên dụng\nTiêu chuẩn mắt lưới xoắn kép\nNhà xưởng 3.000 m2\nCông suất 5 máy đan",
        "applications": "Gia cố mái dốc\nBờ kè sông, kênh\nBảo vệ môi trường\nXây dựng cảnh quan"
    },
    # 12. Màng chống thấm HDPE (Hưng Phú)
    {
        "company": "Công ty CP Vật tư Công trình Hưng Phú",
        "website": "bactham.vn",
        "category": "Màng chống thấm HDPE",
        "name": "Màng chống thấm HDPE HSE",
        "code": "HSE 0.3 – 2.0mm",
        "specs": "HSE 0.3: 0.3mm, kháng thủng 105N\nHSE 0.5: 0.5mm, kháng thủng 176N\nHSE 0.75: 0.75mm, kháng thủng 264N\nHSE 1.0: 1.0mm, kháng thủng 352N\nHSE 1.5: 1.5mm, kháng thủng 540N\nHSE 2.0: 2.0mm, kháng thủng 705N\nKhổ rộng: 8m, Carbon đen 2-3%",
        "features": "Thương hiệu HSE của Aritex\nĐại lý cấp 1\nKho hàng tại Long An\nThi công lắp đặt chuyên nghiệp\nMáy hàn Comet Leister (Thụy Sĩ)",
        "applications": "Lót hồ Biogas\nHồ chứa nước thải\nLót hồ tôm nuôi thủy sản\nChôn lấp rác thải\nBồn chứa xăng dầu"
    },

    # ═══ ARITEX (aritex.com.vn) ═══
    # 13. Vải địa kỹ thuật không dệt ART (Aritex)
    {
        "company": "Công ty CP Vải địa kỹ thuật Việt Nam",
        "website": "aritex.com.vn",
        "category": "Vải địa kỹ thuật không dệt",
        "name": "Vải địa kỹ thuật không dệt ART",
        "code": "ART 7 – ART 28 (Phổ thông)\nART 9D – ART 28D (Loại D)\nART 700G, ART 900G, ART 12A",
        "specs": "Phổ thông: 7-28kN/m, O95: 60-150μm, Khổ 4m\nLoại D: 9.5-28kN/m, O95: 60-180μm, Dày 1.2-3.2mm\nTheo thiết kế: 700G/900G/12A\nCBR: 1200-4500N\nTiêu chuẩn: ASTM D4595, TCVN 8485",
        "features": "Sản xuất tại Việt Nam\nPhương pháp xuyên kim, cán nhiệt\nĐộ bền cao, ổn định kích thước\nĐa dạng: phổ thông, loại D, theo thiết kế",
        "applications": "Phân cách, gia cường, bảo vệ\nLọc và tiêu thoát nước\nXử lý nền đất yếu\nCông trình giao thông, môi trường"
    },
    # 14. Vải địa kỹ thuật dệt GET (Aritex)
    {
        "company": "Công ty CP Vải địa kỹ thuật Việt Nam",
        "website": "aritex.com.vn",
        "category": "Vải địa kỹ thuật dệt",
        "name": "Vải địa kỹ thuật dệt GET",
        "code": "GET 10 – GET 300",
        "specs": "GET 10: ≥100/50 kN/m, CBR ≥4500N\nGET 15: ≥150/50 kN/m, CBR ≥5500N\nGET 20: ≥200/50 kN/m, CBR ≥7000N\nGET 40: ≥400/50 kN/m, CBR ≥12000N\nGET 100: ≥100/100 kN/m, CBR ≥6000N\nGET 200: ≥200/200 kN/m, CBR ≥15000N\nGET 300: ≥300/300 kN/m, CBR ≥18000N\nĐộ giãn: ≤15%, UV >70%, Khổ: 3.5/5.4m",
        "features": "Sản xuất từ PP/PET chất lượng cao\nCường lực chịu kéo cao\nKhả năng lọc và thoát nước tốt\nKháng UV >70%",
        "applications": "Gia cố nền đất yếu\nLọc và thoát nước\nChống xói mòn kè sông, biển\nXây dựng đường bộ, đường sắt, bãi rác"
    },
    # 15. Màng chống thấm HDPE (Aritex)
    {
        "company": "Công ty CP Vải địa kỹ thuật Việt Nam",
        "website": "aritex.com.vn",
        "category": "Màng chống thấm HDPE",
        "name": "Màng chống thấm HDPE HSE",
        "code": "HSE 0.3 – HSE 2.0",
        "specs": "HSE 0.3: 0.3mm, kéo 8kN/m, giãn 600%\nHSE 0.5: 0.5mm, kéo 14kN/m, giãn 700%\nHSE 0.75: 0.75mm, kéo 21kN/m, giãn 700%\nHSE 1.0: 1.0mm, kéo 28kN/m, giãn 700%\nHSE 1.5: 1.5mm, kéo 43kN/m, giãn 700%\nHSE 2.0: 2.0mm, kéo 57kN/m, giãn 700%\nCarbon đen: 2-3%, Khổ: 7-8m\nTuổi thọ >25 năm",
        "features": "Nhà sản xuất HDPE đầu tiên & lớn nhất VN\nSản xuất từ hạt nhựa PP nguyên sinh\nKháng UV, kháng hóa chất\nTuổi thọ >100 năm (lý thuyết)\nKhổ rộng 8m, cắt theo yêu cầu",
        "applications": "Lót hồ nuôi thủy sản\nChôn lấp rác thải\nHồ biogas\nXử lý môi trường ô nhiễm dioxin\nBể chứa xăng dầu"
    },
    # 16. Bấc thấm (Aritex)
    {
        "company": "Công ty CP Vải địa kỹ thuật Việt Nam",
        "website": "aritex.com.vn",
        "category": "Bấc thấm",
        "name": "Bấc thấm đứng & ngang",
        "code": "VID 75, RID 75, RID 4.0, VID 4.5 (đứng)\nRID 200, RID 300 (ngang)",
        "specs": "Đứng: Bề rộng 100mm, Thoát nước ≥80x10-6m3/s (10kPa)\nCường độ kéo >1.6-1.9kN, O95 <0.075mm\nNgang: Bề rộng 200-300mm, Dày 8mm\nThoát nước 80-140x10-6m3/s (100kPa)\nLực chịu nén >250kPa",
        "features": "Lõi PP, vỏ lọc polyester\nChống vi khuẩn, không ăn mòn\nBiến dạng theo địa hình\nDuy trì khả năng thoát nước",
        "applications": "Xử lý gia cố nền đất yếu\nĐường cao tốc, sân bay, đường sắt\nBến cảng, kho xăng dầu\nXử lý môi trường đất ô nhiễm"
    },
    # 17. Màng chống thấm sét tổng hợp (Aritex)
    {
        "company": "Công ty CP Vải địa kỹ thuật Việt Nam",
        "website": "aritex.com.vn",
        "category": "Màng chống thấm sét tổng hợp",
        "name": "Màng chống thấm sét tổng hợp ART Bentonite",
        "code": "ART 3000, ART 4000, ART 4700",
        "specs": "ART 3000: Bentonite >2700g/m2, Thấm ≤5x10-11 m/s\nART 4000: Bentonite >3700g/m2, Thấm ≤3x10-11 m/s\nART 4700: Bentonite >4700g/m2, Thấm ≤5x10-11 m/s\nTrọng lượng vải không dệt >180g/m2\nTrọng lượng vải dệt >110g/m2\nChỉ số trương nở >24ml/2g",
        "features": "Cấu tạo 3 lớp: vải + Bentonite + vải\nTrương nở cao khi gặp nước\nHệ số thấm cực thấp\nChịu khấng bóc ≥65N",
        "applications": "Chôn lấp chất thải\nNiêm phong kín phóng xạ, chất độc\nLót ao hồ chứa chất nhiễm\nXử lý môi trường"
    },
    # 18. Ống địa kỹ thuật (Aritex)
    {
        "company": "Công ty CP Vải địa kỹ thuật Việt Nam",
        "website": "aritex.com.vn",
        "category": "Ống địa kỹ thuật",
        "name": "Ống/Túi địa kỹ thuật",
        "code": "Ống 50x11.2m\nTúi 10x4.4m",
        "specs": "Ống: Kéo >135/110 kN/m, Cường may >76kN/m\nTúi: Kéo >100/80 kN/m, Cường may >55kN/m\nKháng thủng: >10000N (ống), >6500N (túi)\nThấm >0.4 s-1, Kích thước lỗ <0.38mm\nĐộ giãn: <25%",
        "features": "Làm từ vải dệt PET/PP cường lực cao\nKháng nước biển, UV, chịu pH\nThiết kế theo dự án\nISO 9001",
        "applications": "Đê phá sóng ngoài khơi\nKè, bờ bao, lõi đê\nXây dựng hạ tầng hàng hải\nỨng phó lũ lụt"
    },
    # 19. Lưới địa kỹ thuật (Aritex)
    {
        "company": "Công ty CP Vải địa kỹ thuật Việt Nam",
        "website": "aritex.com.vn",
        "category": "Lưới địa kỹ thuật",
        "name": "Lưới địa kỹ thuật",
        "code": "Đơn trục / Hai trục / Cốt sợi",
        "specs": "Lưới đơn trục: Chịu lực 1 chiều\nLưới hai trục: Chịu lực 2 chiều\nLưới cốt sợi: Kết hợp vật liệu khác",
        "features": "Gia cố nền đất\nTăng cường ổn định công trình\nKiểm soát xói mòn\nThoát nước hiệu quả",
        "applications": "Gia cố mái dốc\nỔn định nền móng\nKiểm soát xói mòn\nTăng độ liên kết lớp đất"
    },
    # 20. Ô địa kỹ thuật (Aritex)
    {
        "company": "Công ty CP Vải địa kỹ thuật Việt Nam",
        "website": "aritex.com.vn",
        "category": "Ô địa kỹ thuật",
        "name": "Ô địa kỹ thuật Acel (Geocell)",
        "code": "Acel",
        "specs": "Nguyên liệu: HDPE\nCarbon đen ≥1.5% (ASTM D1603)\nTỉ trọng ≥0.94 g/cm3\nDày tấm: 1.5mm\nChịu lực mối hàn ≥1000 N/10cm",
        "features": "Cấu trúc tổ ong liên hợp\nHàn nhiệt tấm HDPE\nGia cường chịu tải\nChống xói mòn bề mặt",
        "applications": "Gia cố mái dốc\nỔn định nền móng\nChống xói mòn kênh mương\nBảo vệ nền đường, bãi"
    },
]

for idx, p in enumerate(products, 1):
    row = idx + 1
    data = [idx, p["company"], p["website"], p["category"], p["name"], p["code"], p["specs"], p["features"], p["applications"]]
    for col, val in enumerate(data, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.alignment = wrap_align
        cell.border = thin_border
        cell.font = Font(name='Arial', size=10)

# Auto-filter
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(products)+1}"
# Freeze top row
ws1.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════
# SHEET 2: THÔNG SỐ KỸ THUẬT CHI TIẾT
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Thông số kỹ thuật")

spec_headers = ["Nhà cung cấp", "Nhóm sản phẩm", "Mã SP", "Cường độ chịu kéo (kN/m)", "Độ giãn dài (%)", "CBR / Kháng thủng (N)", "Hệ số thấm", "Kích thước lỗ O95", "Khổ rộng (m)", "Độ dày (mm)", "Tiêu chuẩn"]
spec_widths = [22, 22, 18, 22, 16, 22, 16, 18, 14, 14, 20]

for i, (h, w) in enumerate(zip(spec_headers, spec_widths), 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(i)].width = w

spec_data = [
    # ARITEX Vải không dệt - Phổ thông
    ["Aritex", "Vải KĐT không dệt", "ART 7", "7.0", "40-75", "1200", "210 L/m2/s", "150 μm", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT không dệt", "ART 9", "9.0", "40-75", "1500", "170 L/m2/s", "120 μm", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT không dệt", "ART 12", "12.0", "40-75", "1900", "140 L/m2/s", "110 μm", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT không dệt", "ART 15", "15.0", "40-75", "2400", "120 L/m2/s", "90 μm", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT không dệt", "ART 20", "20.0", ">50", "2900", "80 L/m2/s", "75 μm", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT không dệt", "ART 25", "25.0", ">50", "4000", "60 L/m2/s", "70 μm", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT không dệt", "ART 28", "28.0", ">50", "4500", "50 L/m2/s", "60 μm", "4", "", "ASTM D4595 / TCVN 8485"],
    # ARITEX Vải không dệt - Loại D
    ["Aritex", "Vải KĐT không dệt (D)", "ART 9D", "9.5", ">40", "1500", "30×10-4 m/s", "180 μm", "", "1.2", "ASTM D4595"],
    ["Aritex", "Vải KĐT không dệt (D)", "ART 15D", "15.0", ">40", "2400", "30×10-4 m/s", "110 μm", "", "1.9", "ASTM D4595"],
    ["Aritex", "Vải KĐT không dệt (D)", "ART 28D", "28.0", ">40", "4500", "30×10-4 m/s", "60 μm", "", "3.2", "ASTM D4595"],
    # ARITEX Vải dệt
    ["Aritex", "Vải KĐT dệt", "GET 10", "≥100/50", "≤15", "≥4500", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT dệt", "GET 15", "≥150/50", "≤15", "≥5500", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT dệt", "GET 20", "≥200/50", "≤15", "≥7000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT dệt", "GET 100", "≥100/100", "≤15", "≥6000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT dệt", "GET 200", "≥200/200", "≤15", "≥15000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Vải KĐT dệt", "GET 300", "≥300/300", "≤15", "≥18000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    # ARITEX Màng HDPE
    ["Aritex", "Màng HDPE", "HSE 0.3", "8", "600", "105 (xuyên thủng)", "", "", "8", "0.3", "ASTM D5199 / D6693"],
    ["Aritex", "Màng HDPE", "HSE 0.5", "14", "700", "176 (xuyên thủng)", "", "", "8", "0.5", "ASTM D5199 / D6693"],
    ["Aritex", "Màng HDPE", "HSE 0.75", "21", "700", "264 (xuyên thủng)", "", "", "8", "0.75", "ASTM D5199 / D6693"],
    ["Aritex", "Màng HDPE", "HSE 1.0", "28", "700", "352 (xuyên thủng)", "", "", "8", "1.0", "ASTM D5199 / D6693"],
    ["Aritex", "Màng HDPE", "HSE 1.5", "43", "700", "540 (xuyên thủng)", "", "", "8", "1.5", "ASTM D5199 / D6693"],
    ["Aritex", "Màng HDPE", "HSE 2.0", "57", "700", "705 (xuyên thủng)", "", "", "7", "2.0", "ASTM D5199 / D6693"],
    # ARITEX Bentonite
    ["Aritex", "Màng chống thấm sét", "ART 3000", "", "", "", "≤5×10-11 m/s", "", "", "", "ASTM D5084"],
    ["Aritex", "Màng chống thấm sét", "ART 4000", "", "", "", "≤3×10-11 m/s", "", "", "", "ASTM D5084"],
    ["Aritex", "Màng chống thấm sét", "ART 4700", "", "", "", "≤5×10-11 m/s", "", "", "", "ASTM D5084"],
    # ARITEX Bấc thấm
    ["Aritex", "Bấc thấm đứng", "VID 75", ">1.6", ">20", "", "≥1.4×10-4 m/s", "<0.075 mm", "100mm", "", "ASTM D4595 / D4491"],
    ["Aritex", "Bấc thấm đứng", "RID 75", ">1.7", ">20", "", "≥1.4×10-4 m/s", "<0.075 mm", "100mm", "", "ASTM D4595 / D4491"],
    ["Aritex", "Bấc thấm ngang", "RID 200", "", "<25", ">250 (chịu nén)", ">1.4×10-4 m/s", "<0.075 mm", "200mm", "8.0", "ASTM D4595 / D4491"],
    ["Aritex", "Bấc thấm ngang", "RID 300", "", "<25", ">250 (chịu nén)", ">1.4×10-4 m/s", "<0.075 mm", "300mm", "8.0", "ASTM D4595 / D4491"],
    # HƯNG PHÚ
    ["Hưng Phú", "Vải KĐT không dệt", "ART 7", "7", "40-75", "1200", "", "150 μm", "4", "", ""],
    ["Hưng Phú", "Vải KĐT không dệt", "ART 12", "12", "40-75", "1900", "", "110 μm", "4", "", ""],
    ["Hưng Phú", "Vải KĐT không dệt", "ART 15", "15", "40-75", "2400", "", "90 μm", "4", "", ""],
    ["Hưng Phú", "Vải KĐT không dệt", "ART 25", "25", ">50", "4000", "", "70 μm", "4", "", ""],
    ["Hưng Phú", "Vải KĐT dệt", "GET 10", "100/50", "≤15", "≥4500", "", "", "", "", ""],
    ["Hưng Phú", "Vải KĐT dệt", "GET 15", "150/50", "≤15", "≥5500", "", "", "", "", ""],
    ["Hưng Phú", "Vải KĐT dệt", "GET 20", "200/50", "≤15", "≥7000", "", "", "", "", ""],
    ["Hưng Phú", "Vải KĐT dệt", "GET 100", "100/100", "≤15", "≥6000", "", "", "", "", ""],
    ["Hưng Phú", "Vải KĐT dệt", "GET 200", "200/200", "≤15", "≥15000", "", "", "", "", ""],
    ["Hưng Phú", "Màng HDPE", "HSE 0.5", "14", "700", "176", "", "", "8", "0.5", "ASTM"],
    ["Hưng Phú", "Màng HDPE", "HSE 1.0", "28", "700", "352", "", "", "8", "1.0", "ASTM"],
    ["Hưng Phú", "Màng HDPE", "HSE 1.5", "43", "700", "540", "", "", "8", "1.5", "ASTM"],
    ["Hưng Phú", "Rọ đá", "Gabion 8x10", "", "", "", "", "Mắt lưới 8×10cm", "", "", "TCVN 2053"],
    ["Hưng Phú", "Rọ đá", "Gabion 6x8", "", "", "", "", "Mắt lưới 6×8cm", "", "", "TCVN 2053"],
    ["Hưng Phú", "Rọ đá", "Gabion 10x12", "", "", "", "", "Mắt lưới 10×12cm", "", "", "TCVN 2053"],
]

for idx, row_data in enumerate(spec_data, 1):
    row = idx + 1
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.alignment = wrap_align
        cell.border = thin_border
        cell.font = Font(name='Arial', size=10)

ws2.auto_filter.ref = f"A1:{get_column_letter(len(spec_headers))}{len(spec_data)+1}"
ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════
# SHEET 3: SO SÁNH NHÀ CUNG CẤP
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("So sánh nhà cung cấp")

comp_headers = ["Nhóm sản phẩm", "Thái Châu (thaichau.vn)", "Hưng Phú (bactham.vn)", "Aritex (aritex.com.vn)"]
comp_widths = [30, 35, 35, 35]

for i, (h, w) in enumerate(zip(comp_headers, comp_widths), 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws3.column_dimensions[get_column_letter(i)].width = w

comp_data = [
    ["Vải địa kỹ thuật không dệt", "APT 7-80kN/m\nPP 100%, xuyên kim\nCông nghệ Hàn Quốc", "ART 7-28kN/m\nPhổ thông & Loại D\nNhà phân phối", "ART 7-28kN/m\nPhổ thông, Loại D, Theo thiết kế\nNhà sản xuất"],
    ["Vải địa kỹ thuật dệt", "DML\n100% Polyester\nDây chuyền tự động", "GET 10-300 kN/m\nPP/PET\nNhà phân phối", "GET 10-300 kN/m\nPP/PET\nNhà sản xuất"],
    ["Màng chống thấm HDPE", "0.3-3.0mm\n>0.94g/cm3\nNhà sản xuất", "HSE 0.3-2.0mm\nThương hiệu Aritex\nĐại lý cấp 1", "HSE 0.3-2.0mm\nNhà sản xuất đầu tiên & lớn nhất VN\nKhổ 8m"],
    ["Bấc thấm", "VID/RID\nĐứng & Ngang\nNhà sản xuất", "Bấc thấm đứng & ngang\nCung cấp dịch vụ", "VID 75, RID 75, RID 4.0\nVID 4.5, RID 200, RID 300\nNhà sản xuất"],
    ["Màng chống thấm sét (GCL)", "APT GCL\nDệt xuyên kim\nNhà sản xuất", "Không có thông tin", "ART 3000/4000/4700\n3 lớp: vải + Bentonite + vải\nNhà sản xuất"],
    ["Lưới địa kỹ thuật", "PP, PET, Thủy tinh\nĐơn trục, đa trục\nNhà sản xuất", "Không có thông tin", "Đơn trục, hai trục, cốt sợi\nGia cố nền, mái dốc"],
    ["Ô địa kỹ thuật (Geocell)", "HDPE hàn nhiệt\nGia cố mái dốc\nNhà sản xuất", "Không có thông tin", "Acel - HDPE\nCarbon đen ≥1.5%\nTỉ trọng ≥0.94"],
    ["Rọ đá / Thảm đá", "Không có thông tin", "Gabion\nMắt lưới 6-12cm\nMạ kẽm / PVC\nNhà sản xuất (120 Tấn/tháng)", "Không có thông tin"],
    ["Ống địa kỹ thuật (Geotube)", "PET/PP cường lực cao\nKháng nước biển, UV\nNhà sản xuất", "Không có thông tin", "50×11.2m (ống)\n10×4.4m (túi)\nISO 9001"],
    ["Túi vải trồng cây", "Không có thông tin", "Túi có quai xách\nT20-T80\nCombo 20 cái", "Không có thông tin"],
]

for idx, row_data in enumerate(comp_data, 1):
    row = idx + 1
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.alignment = wrap_align
        cell.border = thin_border
        cell.font = Font(name='Arial', size=10)

ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════
output_path = "san_pham_dia_ky_thuat.xlsx"
wb.save(output_path)
print(f"✅ Đã tạo file: {output_path}")
print(f"   - Sheet 1: Tổng hợp sản phẩm ({len(products)} sản phẩm)")
print(f"   - Sheet 2: Thông số kỹ thuật ({len(spec_data)} dòng)")
print(f"   - Sheet 3: So sánh nhà cung cấp ({len(comp_data)} nhóm)")
