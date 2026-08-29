#!/usr/bin/env python3
"""
Scrape and compile product data from 3 geosynthetic websites into Excel.
Sites: thaichau.vn, bactham.vn, aritex.com.vn
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def write_rows(ws, data):
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap_align
            cell.border = thin_border
            cell.font = Font(name='Arial', size=10)

# ═══════════════════════════════════════════════════════
# SHEET 1: ALL PRODUCTS
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Products"

h1 = ["No.", "Supplier", "Website", "Category", "Product Name", "Product Code", "Specifications", "Key Features", "Applications"]
w1 = [6, 28, 22, 28, 38, 22, 60, 48, 48]
style_header(ws1, h1, w1)

products = [
    # ═══ THAI CHAU ═══
    ["Thai Chau", "thaichau.vn", "Vai dia ky thuat khong det", "Vai dia ky thuat khong det APT", "APT 7 - APT 80",
     "Cuong do chiu keo: 7kN - 80kN/m (100g/m2 - 1200g/m2)\nSan xuat 100% PP, xuyen kim, phu gia UV\nDay chuyen tu dong, cong nghe Han Quoc",
     "Da dang cuong luc, tinh on dinh cao\nGia hop ly voi chat luong quoc te\nNguyen lieu PP nhap khau\nHo tro ky thuat chuyen mon",
     "Phan cach, loc, tieu thoat nuoc\nBao ve, gia cuong nen dat yeu\nGia cau cau duong, san bay\nBao ve mang chong tham"],

    ["Thai Chau", "thaichau.vn", "Vai dia ky thuat det", "Vai dia ky thuat det DML", "DML",
     "San xuat tai nha may APT\nDay chuyen tu dong, tieu chuan Han Quoc\n100% nguyen lieu Polyester",
     "Cuong luc chiu keo cao\nKha nang loc va thoat nuoc tot\nDo ben moi truong cao",
     "Gia cuong nen dat yeu\nLoc va thoat nuoc\nChong xi mon ke song, bien"],

    ["Thai Chau", "thaichau.vn", "Bac tham", "Bac tham PVD APT", "VID / RID",
     "Bac tham dung va bac tham ngang\nCau tao: Loi PP/PET + Vo polyester\nCo the dong xuong do sau >40m",
     "Nguyen lieu nhap khau chat luong cao\nChong mai mon tot\nGia hop ly chat luong quoc te\nRut ngan thoi gian thi cong",
     "Gia cuong nen dat yeu\nOn dinh nen cau duong, san bay\nXu ly moi truong dat o nhiem"],

    ["Thai Chau", "thaichau.vn", "Mang chong tham Bentonite", "Mang chong tham Bentonite GCL", "APT GCL",
     "Cong nghe det xuyen kim\nLop Bentonite kep giua 2 vai dia ky thuat\nHe so tham thap hon 10-100 lan so voi dat set day 60-90cm",
     "De thi cong vung kho khan\nNgan o nhiem moi truong\nChat luong kiem soat dong nhat\nTiet kiem van tai",
     "Lot day bai chon lap rac\nLot ao ho chat nhiem bat\nChong tham de dap, kenh muong\nLot day khu vui choi, san golf"],

    ["Thai Chau", "thaichau.vn", "Luoi dia ky thuat", "Luoi dia ky thuat APT", "PP / PET / Thuy tinh",
     "Luoi PP, luoi PET, luoi cot soi thuy tinh\nLuoi hinh chu nhat, luoi tam giac\nDa truc: don truc, da truc",
     "Da dang san pham luoi\nCuong do chiu keo cao, bien dang thap\nBen voi tac dong ly hoa moi truong\nTuoi tho >100 nam",
     "Tuong chan trong luc (cao toi 17m)\nMai doc (cao toi 50m)\nDuong dan dau cau\nCai tao nang cap duong cao toc, san bay"],

    ["Thai Chau", "thaichau.vn", "O dia ky thuat (Geocell)", "O dia ky thuat Geocell APT", "Geocell",
     "Tam HDPE lien ke han nhiet\nSau khi do dat/da soi tao ket cau gia cuong",
     "Gia cuong nen dat\nChong xi mon mai doc\nCau truc to ong linh hoat",
     "Gia cuong mai doc\nOn dinh nen duong\nChong xi mon be mat"],

    ["Thai Chau", "thaichau.vn", "Mang chong tham HDPE", "Mang chong tham HDPE", "HDPE 0.3 - 3.0mm",
     "Tu hat nhua HDPE mat do cao (>0.94g/cm3)\nDo day: 0.3mm - 3.0mm\nCo phu gia chong lao hoa, UV, khang hoa chat\nTuoi tho thiet ke >50 nam",
     "Tinh tro ly, do ben cao\nTiet kiem chi phi van tai\nChat luong kiem soat dong nhat\nKhong gay doc hai moi truong",
     "Bai xu ly rac\nLot ho nuoi thuy san\nMang phu ho biogas\nChong tham de dap, san golf\nLot be chua xang dau"],

    ["Thai Chau", "thaichau.vn", "Ong dia ky thuat (Geotube)", "Ong dia ky thuat Geotube", "Geotube",
     "Vai dia ky thuat det cuong luc cao\nPolyester hoac Polypropylene\nKhang nuoc bien, UV, chiu pH",
     "Chiu moi truong nuoc bien\nKhang tia cuc tim\nCau truc mem mai theo dia hinh",
     "De pha song ngoai khai\nKe, bao bao, loi de\nXay dung ha tang hang hai\nUng pho lu lut"],

    # ═══ BACTHAM (HUNG PHU) ═══
    ["Hung Phu", "bactham.vn", "Vai dia ky thuat khong det", "Vai dia ky thuat khong det ART", "ART 7 - ART 28",
     "ART 7: 7kN/m, CBR 1200N, O95 150um\nART 9: 9kN/m, CBR 1500N, O95 120um\nART 12: 12kN/m, CBR 1900N, O95 110um\nART 15: 15kN/m, CBR 2400N, O95 90um\nART 25: 25kN/m, CBR 4000N, O95 70um\nKho rong: 4m",
     "Da dang cuong luc 7-28kN/m\nChiu keo, xuyen thung cao\nChong mai mon tot\nGia thanh hop ly",
     "Gia cuong nen dat yeu\nLoc va tieu thoat nuoc\nPhan cach lop vat lieu\nBao ve mai doc"],

    ["Hung Phu", "bactham.vn", "Vai dia ky thuat det", "Vai dia ky thuat det GET", "GET 10 - GET 300",
     "GET 10: 100/50 kN/m, CBR >=4500N\nGET 15: 150/50 kN/m, CBR >=5500N\nGET 20: 200/50 kN/m, CBR >=7000N\nGET 100: 100/100 kN/m, CBR >=6000N\nGET 200: 200/200 kN/m, CBR >=15000N\nGET 300: 300/300 kN/m, CBR >=18000N\nDo gian: <=15%, UV >70%",
     "Cuong luc chiu keo cao\nSuat dan hoi cao\nKha nang loc thoat nuot tot\nKhang UV >70%",
     "Gia cuong nen dat yeu\nLoc va thoat nuoc\nChong xi mon ke song, bien\nXay dung duong bo, duong sat"],

    ["Hung Phu", "bactham.vn", "Ro da - Tham da", "Ro da (Gabion)", "Gabion",
     "Mat luoi: 8x10cm, 6x8cm, 10x12cm\nDay dan: Ma kem nhe (TCVN 2053-1993)\nDay dan: Ma kem boc PVC (TCVN 10335-2014)\nMa kem nang: 220-280 g/m2\nCong suat: 120 Tan/thang",
     "Dan tren may chuyen dung\nTieu chuan mat luoi xoan kep\nNha xuong 3.000 m2\nCong suat 5 may dan",
     "Gia cuong mai doc\nBo ke song, kenh\nBao ve moi truong\nXay dung canh quan"],

    ["Hung Phu", "bactham.vn", "Mang chong tham HDPE", "Mang chong tham HDPE HSE", "HSE 0.3 - 2.0mm",
     "HSE 0.3: 0.3mm, khang thung 105N\nHSE 0.5: 0.5mm, khang thung 176N\nHSE 0.75: 0.75mm, khang thung 264N\nHSE 1.0: 1.0mm, khang thung 352N\nHSE 1.5: 1.5mm, khang thung 540N\nHSE 2.0: 2.0mm, khang thung 705N\nKho rong: 8m, Carbon den 2-3%",
     "Thuong hieu HSE cua Aritex\nDai ly cap 1\nKho hang tai Long An\nThi cong lap dat chuyen nghiep\nMay han Comet Leister (Thuy Si)",
     "Lot ho Biogas\nHo chua nuoc thai\nLot ho tom nuoi thuy san\nChon lap rac thai\nBe chua xang dau"],

    # ═══ ARITEX ═══
    ["Aritex", "aritex.com.vn", "Vai dia ky thuat khong det", "Vai KDT khong det ART", "ART 7 - ART 28",
     "Pho thong: 7-28kN/m, O95: 60-150um, Kho 4m\nLoai D: 9.5-28kN/m, O95: 60-180um, Day 1.2-3.2mm\nTheo thiet ke: 700G/900G/12A\nCBR: 1200-4500N\nTieu chuan: ASTM D4595, TCVN 8485",
     "San xuat tai Viet Nam\nPhuong phap xuyen kim, can nhiet\nDo ben cao, on dinh kich thuoc\nDa dang: pho thong, loai D, theo thiet ke",
     "Phan cach, gia cuong, bao ve\nLoc va tieu thoat nuoc\nXu ly nen dat yeu\nCong trinh giao thong, moi truong"],

    ["Aritex", "aritex.com.vn", "Vai dia ky thuat det", "Vai KDT det GET", "GET 10 - GET 300",
     "GET 10: >=100/50 kN/m, CBR >=4500N\nGET 15: >=150/50 kN/m, CBR >=5500N\nGET 20: >=200/50 kN/m, CBR >=7000N\nGET 40: >=400/50 kN/m, CBR >=12000N\nGET 100: >=100/100 kN/m, CBR >=6000N\nGET 200: >=200/200 kN/m, CBR >=15000N\nGET 300: >=300/300 kN/m, CBR >=18000N\nDo gian: <=15%, UV >70%, Kho: 3.5/5.4m",
     "San xuat tu PP/PET chat luong cao\nCuong luc chiu keo cao\nKha nang loc va thoat nuoc tot\nKhang UV >70%",
     "Gia cuong nen dat yeu\nLoc va thoat nuoc\nChong xi mon ke song, bien\nXay dung duong bo, duong sat, bai rac"],

    ["Aritex", "aritex.com.vn", "Mang chong tham HDPE", "Mang chong tham HDPE HSE", "HSE 0.3 - HSE 2.0",
     "HSE 0.3: 0.3mm, keo 8kN/m, gian 600%\nHSE 0.5: 0.5mm, keo 14kN/m, gian 700%\nHSE 0.75: 0.75mm, keo 21kN/m, gian 700%\nHSE 1.0: 1.0mm, keo 28kN/m, gian 700%\nHSE 1.5: 1.5mm, keo 43kN/m, gian 700%\nHSE 2.0: 2.0mm, keo 57kN/m, gian 700%\nCarbon den: 2-3%, Kho: 7-8m\nTuoi tho >25 nam",
     "Nha san xuat HDPE dau tien & lon nhat VN\nSan xuat tu hat nhua PP nguyen sinh\nKhang UV, khang hoa chat\nTuoi tho >100 nam (ly thuyet)\nKho rong 8m, cat theo yeu cau",
     "Lot ho nuoi thuy san\nChon lap rac thai\nHo biogas\nXu ly moi truong o nhiem dioxin\nBe chua xang dau"],

    ["Aritex", "aritex.com.vn", "Bac tham", "Bac tham dung & ngang", "VID 75, RID 75, RID 200, RID 300",
     "Dung: Be rong 100mm, Thoat nuoc >=80x10-6m3/s (10kPa)\nCuong do keo >1.6-1.9kN, O95 <0.075mm\nNgang: Be rong 200-300mm, Day 8mm\nThoat nuoc 80-140x10-6m3/s (100kPa)\nLuc chiu nen >250kPa",
     "Loi PP, vo loc polyester\nChong vi khong, khong an mon\nBien dang theo dia hinh\nDuy tri kha nang thoat nuoc",
     "Xu ly gia cuong nen dat yeu\nDuong cao toc, san bay, duong sat\nBen cang, kho xang dau\nXu ly moi truong dat o nhiem"],

    ["Aritex", "aritex.com.vn", "Mang chong tham set tong hop", "Mang chong tham set ART Bentonite", "ART 3000, ART 4000, ART 4700",
     "ART 3000: Bentonite >2700g/m2, Tham <=5x10-11 m/s\nART 4000: Bentonite >3700g/m2, Tham <=3x10-11 m/s\nART 4700: Bentonite >4700g/m2, Tham <=5x10-11 m/s\nTrong luong vai khong det >180g/m2\nTrong luong vai det >110g/m2\nChi so truong no >24ml/2g",
     "Cau tao 3 lop: vai + Bentonite + vai\nTruong no cao khi gap nuoc\nHe so tham cuc thap\nChiu khac boc >=65N",
     "Chon lap chat thai\nNiem phong kinh phong xa, chat doc\nLot ao ho chat nhiem\nXu ly moi truong"],

    ["Aritex", "aritex.com.vn", "Ong/Tui dia ky thuat", "Ong/Tui dia ky thuat", "Ong 50x11.2m / Tui 10x4.4m",
     "Ong: Keo >135/110 kN/m, Cuong may >76kN/m\nTui: Keo >100/80 kN/m, Cuong may >55kN/m\nKhang thung: >10000N (ong), >6500N (tui)\nTham >0.4 s-1, Kich thuoc lo <0.38mm\nDo gian: <25%",
     "Lam tu vai det PET/PP cuong luc cao\nKhang nuoc bien, UV, chiu pH\nThiet ke theo du an\nISO 9001",
     "De pha song ngoai khai\nKe, bao bao, loi de\nXay dung ha tang hang hai\nUng pho lu lut"],

    ["Aritex", "aritex.com.vn", "Luoi dia ky thuat", "Luoi dia ky thuat", "Don truc / Hai truc / Cot soi",
     "Luoi don truc: Chiu luc 1 chieu\nLuoi hai truc: Chiu luc 2 chieu\nLuoi cot soi: Ket hop vat lieu khac",
     "Gia cuong nen dat\nTang cuong on dinh cong trinh\nKiem soat xi mon\nThoat nuoc hieu qua",
     "Gia cuong mai doc\nOn dinh nen mong\nKiem soat xi mon\nTang do lien ket lop dat"],

    ["Aritex", "aritex.com.vn", "O dia ky thuat", "O dia ky thuat Acel (Geocell)", "Acel",
     "Nguyen lieu: HDPE\nCarbon den >=1.5% (ASTM D1603)\nTi trong >=0.94 g/cm3\nDay tam: 1.5mm\nChiu luc moi han >=1000 N/10cm",
     "Cau truc to ong lien hop\nHan nhiet tam HDPE\nGia cuong chiu tai\nChong xi mon be mat",
     "Gia cuong mai doc\nOn dinh nen mong\nChong xi mon kenh muong\nBao ve nen duong, bai"],
]

write_rows(ws1, products)
ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}{len(products)+1}"

# ═══════════════════════════════════════════════════════
# SHEET 2: SPECS
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Specifications")

h2 = ["Supplier", "Category", "Code", "Tensile (kN/m)", "Elongation (%)", "CBR/Puncture (N)", "Permeability", "Opening O95", "Width (m)", "Thickness (mm)", "Standard"]
w2 = [18, 25, 18, 20, 16, 22, 18, 18, 14, 14, 22]
style_header(ws2, h2, w2)

specs = [
    ["Aritex", "Nonwoven Geotextile", "ART 7", "7.0", "40-75", "1200", "210 L/m2/s", "150 um", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Nonwoven Geotextile", "ART 9", "9.0", "40-75", "1500", "170 L/m2/s", "120 um", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Nonwoven Geotextile", "ART 12", "12.0", "40-75", "1900", "140 L/m2/s", "110 um", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Nonwoven Geotextile", "ART 15", "15.0", "40-75", "2400", "120 L/m2/s", "90 um", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Nonwoven Geotextile", "ART 20", "20.0", ">50", "2900", "80 L/m2/s", "75 um", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Nonwoven Geotextile", "ART 25", "25.0", ">50", "4000", "60 L/m2/s", "70 um", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Nonwoven Geotextile", "ART 28", "28.0", ">50", "4500", "50 L/m2/s", "60 um", "4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Nonwoven Geotextile (D)", "ART 9D", "9.5", ">40", "1500", "30x10-4 m/s", "180 um", "", "1.2", "ASTM D4595"],
    ["Aritex", "Nonwoven Geotextile (D)", "ART 15D", "15.0", ">40", "2400", "30x10-4 m/s", "110 um", "", "1.9", "ASTM D4595"],
    ["Aritex", "Nonwoven Geotextile (D)", "ART 28D", "28.0", ">40", "4500", "30x10-4 m/s", "60 um", "", "3.2", "ASTM D4595"],
    ["Aritex", "Woven Geotextile", "GET 10", ">=100/50", "<=15", ">=4500", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Woven Geotextile", "GET 15", ">=150/50", "<=15", ">=5500", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Woven Geotextile", "GET 20", ">=200/50", "<=15", ">=7000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Woven Geotextile", "GET 100", ">=100/100", "<=15", ">=6000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Woven Geotextile", "GET 200", ">=200/200", "<=15", ">=15000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "Woven Geotextile", "GET 300", ">=300/300", "<=15", ">=18000", "0.02-0.6 s-1", "0.075-0.34 mm", "3.5 / 5.4", "", "ASTM D4595 / TCVN 8485"],
    ["Aritex", "HDPE Liner", "HSE 0.3", "8", "600", "105 (puncture)", "", "", "8", "0.3", "ASTM D5199 / D6693"],
    ["Aritex", "HDPE Liner", "HSE 0.5", "14", "700", "176 (puncture)", "", "", "8", "0.5", "ASTM D5199 / D6693"],
    ["Aritex", "HDPE Liner", "HSE 0.75", "21", "700", "264 (puncture)", "", "", "8", "0.75", "ASTM D5199 / D6693"],
    ["Aritex", "HDPE Liner", "HSE 1.0", "28", "700", "352 (puncture)", "", "", "8", "1.0", "ASTM D5199 / D6693"],
    ["Aritex", "HDPE Liner", "HSE 1.5", "43", "700", "540 (puncture)", "", "", "8", "1.5", "ASTM D5199 / D6693"],
    ["Aritex", "HDPE Liner", "HSE 2.0", "57", "700", "705 (puncture)", "", "", "7", "2.0", "ASTM D5199 / D6693"],
    ["Aritex", "GCL Bentonite", "ART 3000", "", "", "", "<=5x10-11 m/s", "", "", "", "ASTM D5084"],
    ["Aritex", "GCL Bentonite", "ART 4000", "", "", "", "<=3x10-11 m/s", "", "", "", "ASTM D5084"],
    ["Aritex", "GCL Bentonite", "ART 4700", "", "", "", "<=5x10-11 m/s", "", "", "", "ASTM D5084"],
    ["Aritex", "Wick Drain (Vertical)", "VID 75", ">1.6", ">20", "", ">=1.4x10-4 m/s", "<0.075 mm", "100mm", "", "ASTM D4595 / D4491"],
    ["Aritex", "Wick Drain (Vertical)", "RID 75", ">1.7", ">20", "", ">=1.4x10-4 m/s", "<0.075 mm", "100mm", "", "ASTM D4595 / D4491"],
    ["Aritex", "Wick Drain (Horizontal)", "RID 200", "", "<25", ">250 (compression)", ">1.4x10-4 m/s", "<0.075 mm", "200mm", "8.0", "ASTM D4595 / D4491"],
    ["Aritex", "Wick Drain (Horizontal)", "RID 300", "", "<25", ">250 (compression)", ">1.4x10-4 m/s", "<0.075 mm", "300mm", "8.0", "ASTM D4595 / D4491"],
    ["Hung Phu", "Nonwoven Geotextile", "ART 7", "7", "40-75", "1200", "", "150 um", "4", "", ""],
    ["Hung Phu", "Nonwoven Geotextile", "ART 12", "12", "40-75", "1900", "", "110 um", "4", "", ""],
    ["Hung Phu", "Nonwoven Geotextile", "ART 15", "15", "40-75", "2400", "", "90 um", "4", "", ""],
    ["Hung Phu", "Nonwoven Geotextile", "ART 25", "25", ">50", "4000", "", "70 um", "4", "", ""],
    ["Hung Phu", "Woven Geotextile", "GET 10", "100/50", "<=15", ">=4500", "", "", "", "", ""],
    ["Hung Phu", "Woven Geotextile", "GET 15", "150/50", "<=15", ">=5500", "", "", "", "", ""],
    ["Hung Phu", "Woven Geotextile", "GET 20", "200/50", "<=15", ">=7000", "", "", "", "", ""],
    ["Hung Phu", "Woven Geotextile", "GET 100", "100/100", "<=15", ">=6000", "", "", "", "", ""],
    ["Hung Phu", "Woven Geotextile", "GET 200", "200/200", "<=15", ">=15000", "", "", "", "", ""],
    ["Hung Phu", "HDPE Liner", "HSE 0.5", "14", "700", "176", "", "", "8", "0.5", "ASTM"],
    ["Hung Phu", "HDPE Liner", "HSE 1.0", "28", "700", "352", "", "", "8", "1.0", "ASTM"],
    ["Hung Phu", "HDPE Liner", "HSE 1.5", "43", "700", "540", "", "", "8", "1.5", "ASTM"],
    ["Hung Phu", "Gabion", "Gabion 8x10", "", "", "", "", "Mesh 8x10cm", "", "", "TCVN 2053"],
    ["Hung Phu", "Gabion", "Gabion 6x8", "", "", "", "", "Mesh 6x8cm", "", "", "TCVN 2053"],
    ["Hung Phu", "Gabion", "Gabion 10x12", "", "", "", "", "Mesh 10x12cm", "", "", "TCVN 2053"],
]

write_rows(ws2, specs)
ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{len(specs)+1}"

# ═══════════════════════════════════════════════════════
# SHEET 3: COMPARISON
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("Comparison")

h3 = ["Product Category", "Thai Chau (thaichau.vn)", "Hung Phu (bactham.vn)", "Aritex (aritex.com.vn)"]
w3 = [30, 38, 38, 38]
style_header(ws3, h3, w3)

comp = [
    ["Nonwoven Geotextile", "APT 7-80kN/m\n100% PP, needle-punched\nKorean technology", "ART 7-28kN/m\nStandard & Type D\nDistributor", "ART 7-28kN/m\nStandard, Type D, Custom\nManufacturer"],
    ["Woven Geotextile", "DML\n100% Polyester\nAutomatic line", "GET 10-300 kN/m\nPP/PET\nDistributor", "GET 10-300 kN/m\nPP/PET\nManufacturer"],
    ["HDPE Geomembrane", "0.3-3.0mm\n>0.94g/cm3\nManufacturer", "HSE 0.3-2.0mm\nAritex brand\nAuthorized dealer", "HSE 0.3-2.0mm\n1st & largest VN manufacturer\n8m width"],
    ["Prefabricated Vertical Drain", "VID/RID\nVertical & Horizontal\nManufacturer", "Wick drain\nService provider", "VID 75, RID 75, RID 4.0\nVID 4.5, RID 200, RID 300\nManufacturer"],
    ["Geosynthetic Clay Liner", "APT GCL\nNeedle-punched\nManufacturer", "Not available", "ART 3000/4000/4700\n3-layer: fabric + bentonite\nManufacturer"],
    ["Geogrid", "PP, PET, Fiberglass\nUniaxial, Biaxial\nManufacturer", "Not available", "Uniaxial, Biaxial, Fiberglass\nSoil reinforcement"],
    ["Geocell", "HDPE welded\nSlope protection\nManufacturer", "Not available", "Acel - HDPE\nCarbon black >=1.5%\nDensity >=0.94"],
    ["Gabion / Reno Mattress", "Not available", "Gabion\nMesh 6-12cm\nGalvanized / PVC\nManufacturer (120 T/month)", "Not available"],
    ["Geotextile Tube", "PET/PP high-strength\nMarine, UV resistant\nManufacturer", "Not available", "50x11.2m (tube)\n10x4.4m (bag)\nISO 9001"],
    ["Fabric Planting Bag", "Not available", "With handle\nT20-T80\nCombo 20 pcs", "Not available"],
]

write_rows(ws3, comp)

# ═══════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════
output = "san_pham_dia_ky_thuat.xlsx"
wb.save(output)
print(f"Done: {output}")
print(f"  Sheet 1: {len(products)} products")
print(f"  Sheet 2: {len(specs)} spec rows")
print(f"  Sheet 3: {len(comp)} comparison rows")
