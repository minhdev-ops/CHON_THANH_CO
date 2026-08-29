#!/usr/bin/env python3
"""
Convert san_pham_dia_ky_thuat.xlsx to CSV files for database import.
"""

import openpyxl
import csv
import os

wb = openpyxl.load_workbook('san_pham_dia_ky_thuat.xlsx')

output_dir = 'csv_export'
os.makedirs(output_dir, exist_ok=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    filename = f"{output_dir}/{sheet_name.lower().replace(' ', '_')}.csv"

    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            # Replace None with empty string
            writer.writerow([str(cell) if cell is not None else '' for cell in row])

    print(f"Exported: {filename} ({ws.max_row} rows)")

print(f"\nDone! Files saved to: {output_dir}/")
print("Use utf-8-sig encoding when importing to database.")
